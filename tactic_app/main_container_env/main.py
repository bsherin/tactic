import sys
import re
# noinspection PyUnresolvedReferences
import requests
import copy
import gevent
from gevent import monkey; monkey.patch_all()
from flask import Flask, render_template
from gevent.queue import Queue
import pymongo
import gridfs
import gridfs
import time
import uuid
import cPickle
from bson.binary import Binary
from qworker import QWorker, task_worthy
import datetime
from communication_utils import send_request_to_container
import numpy
import traceback
import zlib
import openpyxl
import cStringIO

INITIAL_LEFT_FRACTION = .69
CHUNK_SIZE = 200
STEP_SIZE = 100


# noinspection PyPep8Naming
class docInfo:
    def __init__(self, name, data_rows, header_list=None, cell_backgrounds=None):

        self.name = name
        self.data_rows = copy.deepcopy(data_rows)  # All the data rows in the doc
        self.current_data_rows = self.data_rows  # The current filtered set of data rows
        self.header_list = header_list
        self.table_spec = {}
        self.start_of_current_chunk = None
        self.is_first_chunk = None
        self.infinite_scroll_required = None
        self.is_last_chunk = None
        if cell_backgrounds is None:
            self.cell_backgrounds = {}
        else:
            self.cell_backgrounds = cell_backgrounds
        self.configure_for_current_data()
        if len(self.data_rows.keys()) > CHUNK_SIZE:
            self.max_table_size = CHUNK_SIZE
        else:
            self.max_table_size = len(self.data_rows.keys())

    def set_background_color(self, row, column_header, color):
        if not str(row) in self.cell_backgrounds:
            self.cell_backgrounds[str(row)] = {}
        self.cell_backgrounds[str(row)][column_header] = color

    @property
    def displayed_background_colors(self):
        result = {}
        sorted_int_keys = sorted([int(key) for key in self.current_data_rows.keys()])
        for i, r in enumerate(sorted_int_keys[self.start_of_current_chunk:(self.start_of_current_chunk + CHUNK_SIZE)]):
            if str(r) in self.cell_backgrounds:
                result[i] = self.cell_backgrounds[str(r)]
        return result

    def configure_for_current_data(self):
        self.start_of_current_chunk = 0
        self.is_first_chunk = True
        if len(self.current_data_rows.keys()) <= CHUNK_SIZE:
            self.infinite_scroll_required = False
            self.is_last_chunk = True
        else:
            self.infinite_scroll_required = True
            self.is_last_chunk = False

    def get_actual_row(self, row_id):
        for i, the_row in enumerate(self.displayed_data_rows):
            if str(row_id) == str(the_row["__id__"]):
                return i
        return None

    def compile_save_dict(self):
        return ({"name": self.name,
                 "data_rows": self.data_rows,
                 "table_spec": self.table_spec,
                 "cell_backgrounds": self.cell_backgrounds,
                 "my_class_for_recreate": "docInfo",
                 "header_list": self.header_list})

    def get_id_from_actual_row(self, actual_row):
        return self.sorted_data_rows[actual_row]["__id__"]

    @property
    def sorted_data_rows(self):
        result = []
        sorted_int_keys = sorted([int(key) for key in self.current_data_rows.keys()])
        for r in sorted_int_keys:
            result.append(self.data_rows[str(r)])
        return result

    @property
    def all_sorted_data_rows(self):
        result = []
        sorted_int_keys = sorted([int(key) for key in self.data_rows.keys()])
        for r in sorted_int_keys:
            result.append(self.data_rows[str(r)])
        return result

    @property
    def displayed_data_rows(self):
        if not self.infinite_scroll_required:
            return self.sorted_data_rows
        result = []
        sorted_int_keys = sorted([int(key) for key in self.current_data_rows.keys()])
        for r in sorted_int_keys[self.start_of_current_chunk:(self.start_of_current_chunk + CHUNK_SIZE)]:
            result.append(self.current_data_rows[str(r)])
        return result

    def advance_to_next_chunk(self):
        if self.is_last_chunk:
            return
        old_start = self.start_of_current_chunk
        self.start_of_current_chunk += STEP_SIZE
        self.is_first_chunk = False
        if (self.start_of_current_chunk + CHUNK_SIZE) >= len(self.current_data_rows):
            self.start_of_current_chunk = len(self.current_data_rows) - CHUNK_SIZE
            self.is_last_chunk = True
        return self.start_of_current_chunk - old_start

    def row_is_visible(self, row_id):
        sorted_int_keys = sorted([int(key) for key in self.current_data_rows.keys()])
        displayed_int_keys = sorted_int_keys[self.start_of_current_chunk:(self.start_of_current_chunk + STEP_SIZE)]
        return int(row_id) in displayed_int_keys

    def move_to_row(self, row_id):
        self.current_data_rows = self.data_rows  # Undo any filtering
        self.configure_for_current_data()
        while not self.is_last_chunk and not(self.row_is_visible(row_id)):
            self.advance_to_next_chunk()

    def go_to_previous_chunk(self):
        if self.is_first_chunk:
            return None
        old_start = self.start_of_current_chunk
        self.start_of_current_chunk = self.start_of_current_chunk - STEP_SIZE
        self.is_last_chunk = False
        if self.start_of_current_chunk <= 0:
            self.start_of_current_chunk = 0
            self.is_first_chunk = True
        return old_start - self.start_of_current_chunk

    @property
    def data_rows_int_keys(self):
        result = {}
        for (key, val) in self.data_rows.items():
            result[int(key)] = val
        return result

    @staticmethod
    def recreate_from_save(save_dict):
        new_instance = docInfo(save_dict["name"],
                               save_dict["data_rows"],
                               save_dict["header_list"],
                               save_dict["cell_backgrounds"])
        new_instance.table_spec = save_dict["table_spec"]
        return new_instance


# noinspection PyPep8Naming
class mainWindow(QWorker):
    save_attrs = ["short_collection_name", "collection_name", "current_tile_id", "tile_sort_list", "left_fraction",
                  "is_shrunk", "doc_dict", "project_name", "loaded_modules", "user_id",
                  "hidden_columns_list", "console_html"]
    update_events = ["CellChange", "CreateColumn", "SearchTable", "SaveTableSpec", "MainClose", "DisplayCreateErrors",
                     "DehighlightTable", "SetCellContent", "RemoveTile", "ColorTextInCell",
                     "FilterTable", "UnfilterTable", "TextSelect", "UpdateSortList", "UpdateLeftFraction",
                     "UpdateTableShrinkState"]

    # noinspection PyUnresolvedReferences
    def __init__(self, app, data_dict):
        QWorker.__init__(self, app, data_dict["megaplex_address"], data_dict["main_id"])
        try:
            client = pymongo.MongoClient(data_dict["mongo_uri"])
            client.server_info()
            # noinspection PyUnresolvedReferences
            self.db = client.tacticdb
            self.fs = gridfs.GridFS(self.db)
        except pymongo.errors.PyMongoError as err:
            sys.exit()

        self.base_figure_url = data_dict["base_figure_url"]
        self.project_collection_name = data_dict["project_collection_name"]

        # These are the main attributes that define a project state

        self.callbacks = {}

        # These are working attributes that will change whenever the project is instantiated.

        self._change_list = []
        self.visible_doc_name = None
        self._pipe_dict = {}
        self.selected_text = ""
        self.recreate_errors = []
        self.project_dict = None
        self.tile_save_results = None
        self.mdata = None

        if "project_name" not in data_dict:
            self.current_tile_id = 0
            self.tile_instances = {}
            self.tile_sort_list = []
            self.left_fraction = INITIAL_LEFT_FRACTION
            self.is_shrunk = False
            self.collection_name = data_dict["collection_name"]
            self.short_collection_name = re.sub("^.*?\.data_collection\.", "", self.collection_name)
            self.project_name = None
            self.console_html = None
            self.user_id = data_dict["user_id"]
            self.doc_dict = self._build_doc_dict()
            self.debug_log("Done with build_doc_dict")
            self.visible_doc_name = self.doc_dict.keys()[0]

    # Communication Methods

    def ask_host(self, msg_type, task_data=None, callback_func=None):
        task_data["main_id"] = self.my_id
        self.post_task("host", msg_type, task_data, callback_func)
        return

    def ask_tile(self, tile_id, msg_type, task_data=None, callback_func=None):
        self.post_task(tile_id, msg_type, task_data, callback_func)
        return

    def emit_table_message(self, message, data=None, callback_func=None):
        if data is None:
            data = {}
        data["table_message"] = message
        self.ask_host("emit_table_message", data, callback_func)
        return

    def distribute_event(self, event_name, data_dict=None, tile_id=None):
        self.debug_log("Entering distribute event with event_name: " + event_name)
        if data_dict is None:
            data_dict = {}
        if tile_id is not None:
            self.ask_tile(tile_id, event_name, data_dict)
        else:
            for tile_id in self.tile_instances.keys():
                self.ask_tile(tile_id, event_name, data_dict)
        if event_name in self.update_events:
            self.post_task(self.my_id, event_name, data_dict)
        self.debug_log("successfully leaving distribute_event in main with with event_name: " + event_name)
        return True

    # Save and load-related methods

    def compile_save_dict(self):
        self.debug_log("entering compile_save_dict in main")
        result = {}
        for attr in self.save_attrs:
            attr_val = getattr(self, attr)
            if hasattr(attr_val, "compile_save_dict"):
                result[attr] = attr_val.compile_save_dict()
            elif (type(attr_val) == dict) and(len(attr_val) > 0) and hasattr(attr_val.values()[0], "compile_save_dict"):
                res = {}
                for (key, val) in attr_val.items():
                    res[key] = val.compile_save_dict()
                result[attr] = res
            else:
                result[attr] = attr_val
        tile_instances = {}
        for tile_id in self.tile_instances.keys():
            tile_save_dict = self.post_and_wait(tile_id, "compile_save_dict")
            tile_instances[tile_id] = tile_save_dict  # tile_id isn't meaningful going forward.
        result["tile_instances"] = tile_instances
        self.debug_log("leaving compile_save_dict in main")
        return result

    def show_um_message(self, message, user_manage_id, timeout=None):
        data = {"message": message, "timeout": timeout, "user_manage_id": user_manage_id}
        self.post_task("host", "show_um_status_message_task", data)

    def clear_um_message(self, user_manage_id):
        data = {"user_manage_id": user_manage_id}
        self.post_task("host", "clear_um_status_message_task", data)

    def show_main_status_message(self, message, timeout=None):
        data = {"message": message, "timeout": timeout, "main_id": self.my_id}
        self.post_task("host", "show_main_status_message", data)

    def clear_main_status_message(self):
        data = {"main_id": self.my_id}
        self.post_task("host", "clear_main_status_message", data)

    @task_worthy
    def do_full_recreation(self, data_dict):
        try:
            tile_info_dict, loaded_modules = self.recreate_from_save(data_dict["project_collection_name"],
                                                                     data_dict["project_name"])
            self.post_and_wait("host", "load_modules", {"loaded_modules": loaded_modules, "user_id": data_dict["user_id"]})
            doc_names = [str(doc_name) for doc_name in self.doc_names]
            self.show_um_message("Creating empty containers", data_dict["user_manage_id"])
            tile_containers = self.post_and_wait("host", "get_empty_tile_containers", {"number": len(tile_info_dict.keys())})
            self.show_um_message("Getting tile code", data_dict["user_manage_id"])
            tile_code_dict = self.post_and_wait("host", "get_tile_code", {"tile_info_dict": tile_info_dict,
                                                                          "user_id": data_dict["user_id"]})
            new_tile_info = {}
            new_tile_keys = tile_containers.keys()
            for i, old_tile_id in enumerate(tile_info_dict.keys()):
                new_id = new_tile_keys[i]
                new_address = tile_containers[new_tile_keys[i]]
                new_tile_info[old_tile_id] = {"new_tile_id": new_id,
                                              "tile_container_address": new_address}
                result = send_request_to_container(new_address, "load_source",
                                                   {"tile_code": tile_code_dict[old_tile_id],
                                                    "megaplex_address": self.megaplex_address}).json()
                if not result["success"]:
                    raise Exception(result["message_string"])
            self.show_um_message("Recreating the tiles", data_dict["user_manage_id"])
            self.tile_save_results = self.recreate_project_tiles(data_dict["list_names"], new_tile_info)
            template_data = {"collection_name": self.collection_name,
                             "project_name": self.project_name,
                             "window_title": self.project_name,
                             "main_id": self.my_id,
                             "doc_names": doc_names,
                             "use_ssl": str(data_dict["use_ssl"]),
                             "console_html": self.console_html,
                             "short_collection_name": self.short_collection_name,
                             "new_tile_info": new_tile_info}

            self.clear_um_message(data_dict["user_manage_id"])
            self.post_task("host", "open_project_window", {"user_manage_id": data_dict["user_manage_id"],
                                                           "template_data": template_data,
                                                           "message": "window-open"})
        except Exception as ex:
            template = "An exception of type {0} occured. Arguments:\n{1!r}\n"
            error_string = template.format(type(ex).__name__, ex.args)
            error_string += traceback.format_exc()
            error_string = "<pre>" + error_string + "</pre>"
            self.post_task("host", "open_error_window", {"user_manage_id": data_dict["user_manage_id"],
                                                         "error_string": error_string})
        return

    def recreate_from_save(self, project_collection_name, project_name):
        self.debug_log("entering recreate_from_save")
        save_dict = self.db[project_collection_name].find_one({"project_name": project_name})
        project_dict = cPickle.loads(zlib.decompress(self.fs.get(save_dict["file_id"]).read()).decode("utf-8", "ignore").encode("ascii"))
        project_dict["metadata"] = save_dict["metadata"]
        self.mdata = save_dict["metadata"]
        for (attr, attr_val) in project_dict.items():
            if str(attr) != "tile_instances":
                try:
                    if type(attr_val) == dict and ("my_class_for_recreate" in attr_val):
                        cls = getattr(sys.modules[__name__], attr_val["my_class_for_recreate"])
                        setattr(self, attr, cls.recreate_from_save(attr_val))
                    elif (type(attr_val) == dict) and (len(attr_val) > 0) and \
                            ("my_class_for_recreate" in attr_val.values()[0]):
                        cls = getattr(sys.modules[__name__], attr_val.values()[0]["my_class_for_recreate"])
                        res = {}
                        for (key, val) in attr_val.items():
                            tinstance = cls.recreate_from_save(val)
                            if tinstance is not None:
                                res[key] = tinstance
                            else:
                                error_messages.append("error creating {}".format(key))
                        setattr(self, attr, res)

                    else:
                        setattr(self, attr, attr_val)
                except TypeError:
                    setattr(self, attr, attr_val)

        for attr in self.save_attrs:
            if attr not in project_dict:
                setattr(self, attr, "")

        tile_info_dict = {}
        for old_tile_id, tile_save_dict in project_dict["tile_instances"].items():
            tile_info_dict[old_tile_id] = tile_save_dict["tile_type"]
        self.project_dict = project_dict
        # self.doc_dict = self._build_doc_dict()
        self.visible_doc_name = self.doc_dict.keys()[0]  # This is necessary for recreating the tiles
        return tile_info_dict, project_dict["loaded_modules"]

    def recreate_project_tiles(self, list_names, new_tile_info):
        self.tile_instances = {}
        tile_results = {}
        for old_tile_id, tile_save_dict in self.project_dict["tile_instances"].items():
            new_tile_id = new_tile_info[old_tile_id]["new_tile_id"]
            new_tile_address = new_tile_info[old_tile_id]["tile_container_address"]
            self.tile_instances[new_tile_id] = new_tile_address
            self.tile_sort_list[self.tile_sort_list.index(old_tile_id)] = new_tile_id
            tile_save_dict = self.project_dict["tile_instances"][old_tile_id]
            tile_save_dict["tile_id"] = new_tile_id
            tile_save_dict["main_id"] = self.my_id
            tile_save_dict["new_base_figure_url"] = self.base_figure_url.replace("tile_id", new_tile_id)
            tresult = send_request_to_container(new_tile_address,
                                                "recreate_from_save",
                                                tile_save_dict,
                                                timeout=60, tries=30)
            tile_result = tresult.json()
            if not tile_result["success"]:
                raise Exception(tile_result["message_string"])
            tile_results[new_tile_id] = tile_result

        for tile in self.tile_sort_list:
            if tile not in self.tile_instances:
                self.tile_sort_list.remove(tile)

        # There's some extra work I have to do once all of the tiles are built.
        # Each tile needs to know the main_id it's associated with.
        # Also I have to build the pipe machinery.
        for tile_id, tile_result in tile_results.items():
            if "exports" in tile_result:
                if len(tile_result["exports"]) > 0:
                    if tile_id not in self._pipe_dict:
                        self._pipe_dict[tile_id] = {}
                    for export in tile_result["exports"]:
                        self._pipe_dict[tile_id][tile_result["tile_name"] + "_" + export] = {
                            "export_name": export,
                            "tile_id": tile_id,
                            "tile_address": self.tile_instances[tile_id],}

        # We have to wait to here to actually render the tiles because
        # the pipe_dict needs to be complete to build the forms.

        form_info = {"current_header_list": self.current_header_list,
                     "pipe_dict": self._pipe_dict,
                     "doc_names": self.doc_names,
                     "list_names": list_names}
        for tile_id, tile_result in tile_results.items():
            tile_result["tile_html"] = self.post_and_wait(tile_id, "render_tile", form_info)["tile_html"]
        return tile_results

    @task_worthy
    def save_new_project(self, data_dict):
        self.debug_log("entering save_new_project")
        # noinspection PyBroadException
        try:
            self.project_name = data_dict["project_name"]
            self.hidden_columns_list = data_dict["hidden_columns_list"]
            self.console_html = data_dict["console_html"]
            tspec_dict = data_dict["tablespec_dict"]
            for (dname, spec) in tspec_dict.items():
                self.doc_dict[dname].table_spec = spec

            self.show_main_status_message("Getting loaded modules")
            self.loaded_modules = self.post_and_wait("host", "get_loaded_user_modules", {"user_id": self.user_id})[
                "loaded_modules"]
            self.loaded_modules = [str(module) for module in self.loaded_modules]
            self.show_main_status_message("compiling save dictionary")
            project_dict = self.compile_save_dict()

            save_dict = {}
            save_dict["metadata"] = self.create_initial_metadata()
            save_dict["project_name"] = project_dict["project_name"]
            self.show_main_status_message("Pickle, convert, compress")
            pdict = cPickle.dumps(project_dict)
            pdict = Binary(zlib.compress(pdict))
            self.show_main_status_message("Writing the data")
            save_dict["file_id"] = self.fs.put(pdict)
            self.mdata = save_dict["metadata"]
            self.db[self.project_collection_name].insert_one(save_dict)
            self.clear_main_status_message()

            return_data = {"project_name": data_dict["project_name"],
                           "success": True,
                           "message_string": "Project Successfully Saved"}

        except Exception as ex:
            self.debug_log("got an error in save_new_project")
            error_string = self.handle_exception(ex, "<pre>Error saving new project</pre>", print_to_console=False)
            return_data = {"success": False, "message_string": error_string}
        return return_data

    @task_worthy
    def update_project(self, data_dict):
        self.debug_log("entering update_project")
        # noinspection PyBroadException
        try:
            self.hidden_columns_list = data_dict["hidden_columns_list"]
            self.console_html = data_dict["console_html"]
            tspec_dict = data_dict["tablespec_dict"]
            for (dname, spec) in tspec_dict.items():
                self.doc_dict[dname].table_spec = spec
            self.show_main_status_message("Getting loaded modules")
            self.loaded_modules = self.post_and_wait("host", "get_loaded_user_modules", {"user_id": self.user_id})[
                "loaded_modules"]
            self.loaded_modules = [str(module) for module in self.loaded_modules]
            self.show_main_status_message("compiling save dictionary")
            project_dict = self.compile_save_dict()
            pname = project_dict["project_name"]
            self.mdata["updated"] = datetime.datetime.today()
            self.show_main_status_message("Pickle, convert, compress")
            pdict = cPickle.dumps(project_dict)
            pdict = Binary(zlib.compress(pdict))
            self.show_main_status_message("Writing the data")
            new_file_id = self.fs.put(pdict)
            save_dict = self.db[self.project_collection_name].find_one({"project_name": pname})
            self.fs.delete(save_dict["file_id"])
            save_dict["project_name"] = pname
            save_dict["metadata"] = self.mdata
            save_dict["file_id"] = new_file_id
            self.db[self.project_collection_name].update_one({"project_name": pname},
                                                             {'$set': save_dict})
            self.clear_main_status_message()
            self.mdata = save_dict["metadata"]
            return_data = {"project_name": pname,
                           "success": True,
                           "message": "Project Successfully Saved"}

        except Exception as ex:
            error_string = self.handle_exception(ex, "Error saving project", print_to_console=False)
            return_data = {"success": False, "message": error_string}
        return return_data

    # utility methods

    def _build_doc_dict(self):
        result = {}
        self.debug_log("building doc_dict with collection: " + self.collection_name)
        the_collection = self.db[self.collection_name]
        for f in the_collection.find():
            fname = f["name"].encode("ascii", "ignore")
            self.debug_log("Got fname " + fname)
            if fname == "__metadata__":
                continue
            if "header_list" in f:
                # Note conversion of unicode filenames to strings
                self.debug_log("got header_list " + str(f["header_list"]))
                result[fname] = docInfo(fname, f["data_rows"], f["header_list"])
            else:
                result[fname] = docInfo(fname, f["data_rows"], [])
        return result

    def _set_row_column_data(self, doc_name, the_id, column_header, new_content):
        self.doc_dict[doc_name].data_rows[str(the_id)][column_header] = new_content
        return

    def print_to_console(self, message_string, force_open=False):
        with self.app.test_request_context():
            pmessage = render_template("log_item.html", log_item=message_string)
        self.emit_table_message("consoleLog", {"message_string": pmessage, "force_open": force_open})
        return {"success": True}

    @property
    def doc_names(self):
        return sorted([str(key) for key in self.doc_dict.keys()])

    def tablespec_dict(self):
        tdict = {}
        for (key, docinfo) in self.doc_dict.items():
            if docinfo.table_spec:  # This will be false if table_spec == {}
                tdict[key] = docinfo.table_spec
        return tdict

    def create_initial_metadata(self):
        mdata = {"datetime": datetime.datetime.today(),
                 "updated": datetime.datetime.today(),
                 "tags": "",
                 "notes": ""}
        return mdata

    def refill_table(self):
        doc = self.doc_dict[self.visible_doc_name]
        self.debug_log("current_data_rows keys " + str(doc.current_data_rows.keys()))
        self.debug_log("displayed_data_rows keys " + str(len(doc.displayed_data_rows)))
        data_object = {"data_rows": doc.displayed_data_rows, "doc_name": self.visible_doc_name,
                       "background_colors": doc.displayed_background_colors,
                       "is_first_chunk": doc.is_first_chunk, "is_last_chunk": doc.is_last_chunk}
        self.emit_table_message("refill_table", data_object)

    @property
    def tile_ids(self):
        return self.tile_instances

    @property
    def current_header_list(self):
        dinfo = self.doc_dict[self.visible_doc_name]
        return dinfo.header_list

    def get_list_names(self):
        list_names = self.post_and_wait("host", "get_list_names", {"user_id": self.user_id})
        return list_names

    def _delete_tile_instance(self, tile_id):
        send_request_to_container(self.tile_instances[tile_id], "kill_me")
        del self.tile_instances[tile_id]
        self.tile_sort_list.remove(tile_id)
        form_info = {"current_header_list": self.current_header_list,
                     "pipe_dict": self._pipe_dict,
                     "doc_names": self.doc_names,
                     "list_names": self.get_list_names()["list_names"]}
        if tile_id in self._pipe_dict:
            del self._pipe_dict[tile_id]
            for tid in self.tile_instances.keys():
                self.post_task(tid, "RebuildTileForms", form_info)
        self.ask_host("delete_container", {"container_id": tile_id})
        return

    def get_current_pipe_list(self):
        pipe_list = []
        for tile_entry in self._pipe_dict.values():
            pipe_list += tile_entry.keys()
        return pipe_list

    def handle_exception(self, ex, special_string=None, print_to_console=True):
        if special_string is None:
            template = "An exception of type {0} occured. Arguments:\n{1!r}\n"
        else:
            template = special_string + "\n" + "An exception of type {0} occurred. Arguments:\n{1!r}\n"
        error_string = template.format(type(ex).__name__, ex.args)
        error_string += traceback.format_exc()
        error_string = "<pre>" + error_string + "</pre>"
        self.debug_log(error_string)
        if print_to_console:
            self.print_to_console(error_string, force_open=True)
        return

    def highlight_table_text(self, txt):
        row_index = 0
        dinfo = self.doc_dict[self.visible_doc_name]
        for the_row in dinfo.displayed_data_rows:
            for cheader in dinfo.header_list:
                cdata = the_row[cheader]
                if cdata is None:
                    continue
                if str(txt).lower() in str(cdata).lower():
                    self.emit_table_message("highlightTxtInCell",
                                            {"row_index": row_index, "column_header": cheader, "text_to_find": txt})
            row_index += 1

    @staticmethod
    def txt_in_dict(txt, d):
        for val in d.values():
            try:
                if str(txt).lower() in str(val).lower():
                    return True
            except UnicodeEncodeError:
                continue
        return False

    # Task Worthy methods. These are eligible to be the recipient of posted tasks.
    @task_worthy
    def create_tile(self, data_dict):
        self.debug_log("entering create_tile in main.py")
        tile_container_id = data_dict["tile_id"]
        self.tile_instances[tile_container_id] = data_dict["tile_address"]
        tile_name = data_dict["tile_name"]
        data_dict["user_id"] = self.user_id
        data_dict["base_figure_url"] = self.base_figure_url.replace("tile_id", tile_container_id)
        data_dict["main_id"] = self.my_id
        data_dict["megaplex_address"] = self.megaplex_address
        form_info = {"current_header_list": self.current_header_list,
                     "pipe_dict": self._pipe_dict,
                     "doc_names": self.doc_names,
                     "list_names": data_dict["list_names"]}
        # data_dict["form_info"] = form_info
        tile_address = data_dict["tile_address"]
        result = send_request_to_container(tile_address, "load_source", data_dict).json()
        if not result["success"]:
            raise Exception(result["message_string"])

        instantiate_result = send_request_to_container(tile_address, "instantiate_tile_class", data_dict).json()
        if not instantiate_result["success"]:
            raise Exception(instantiate_result["message_string"])

        exports = instantiate_result["exports"]
        if len(exports) > 0:
            if tile_container_id not in self._pipe_dict:
                self._pipe_dict[tile_container_id] = {}
            for export in exports:
                self._pipe_dict[tile_container_id][tile_name + "_" + export] = {
                    "export_name": export,
                    "tile_id": tile_container_id,
                    "tile_address": tile_address}

        form_html = self.post_and_wait(tile_container_id, "create_form_html", form_info)["form_html"]
        for tid in self.tile_instances.keys():
            if not tid == tile_container_id:
                self.post_task(tile_container_id, "RebuildTileForms", form_info)
        self.tile_sort_list.append(tile_container_id)
        self.current_tile_id += 1
        self.debug_log("leaving create_tile in main.py")
        return {"success": True, "html": form_html}

    @task_worthy
    def get_func(self, data_dict):  # tactic_todo this is too powerful for tile API
        func_name = data_dict["func"]
        args = data_dict["args"]
        val = getattr(self, func_name)(*args)
        return {"val": val}

    @task_worthy
    def get_column_data(self, data):
        self.debug_log("entering get_column_data")
        result = []
        ddata = copy.copy(data)
        for doc_name in self.doc_dict.keys():
            ddata["doc_name"] = doc_name
            result = result + self.get_column_data_for_doc(ddata)
        self.debug_log("leaving get_column_data")
        return result

    @task_worthy
    def get_document_data(self, data):
        self.debug_log("entering get_document_data")
        doc_name = data["document_name"]
        return self.doc_dict[doc_name].data_rows_int_keys

    @task_worthy
    def get_document_data_as_list(self, data):
        self.debug_log("entering get_document_data_as_list")
        doc_name = data["document_name"]
        data_list = self.doc_dict[doc_name].all_sorted_data_rows
        return {"data_list": data_list}

    @task_worthy
    def get_column_names(self, data):
        self.debug_log("entering get_column_names")
        doc_name = data["document_name"]
        header_list = self.doc_dict[doc_name].header_list
        return {"header_list": header_list}

    @task_worthy
    def get_number_rows(self, data):
        self.debug_log("entering get_column_names")
        doc_name = data["document_name"]
        nrows = len(self.doc_dict[doc_name].data_rows.keys())
        return {"number_rows": nrows}

    @task_worthy
    def get_row(self, data):
        self.debug_log("entering get_column_names")
        doc_name = data["document_name"]
        row_num = data["row_number"]
        the_row = self.doc_dict[doc_name].all_sorted_data_rows[int(row_num)]
        return the_row

    @task_worthy
    def get_cell(self, data):
        self.debug_log("entering get_column_names")
        doc_name = data["document_name"]
        row_num = data["row_number"]
        column_name = data["column_name"]
        the_cell = self.doc_dict[doc_name].all_sorted_data_rows[int(row_num)][column_name]
        return {"the_cell": the_cell}

    @task_worthy
    def get_column_data_for_doc(self, data):
        self.debug_log("entering get_column_data_for_doc in main.py")
        column_header = data["column_name"]
        doc_name = data["doc_name"]
        the_rows = self.doc_dict[doc_name].all_sorted_data_rows
        result = []
        for the_row in the_rows:
            result.append(the_row[column_header])
        self.debug_log("leaving get_column_data_for_doc in main.py")
        return result

    @task_worthy
    def CellChange(self, data):
        self.debug_log("Entering CellChange in main.")
        self._set_row_column_data(data["doc_name"], data["id"], data["column_header"], data["new_content"])
        self._change_list.append(data["id"])
        return None

    @task_worthy
    def set_visible_doc(self, data):
        self.debug_log("entering save visible_doc on main")
        doc_name = data["doc_name"]
        self.visible_doc_name = doc_name
        return {"success": True}

    @task_worthy
    def print_to_console_event(self, data):
        self.print_to_console(data["print_string"], force_open=True)
        return None

    @task_worthy
    def get_property(self, data_dict):
        self.debug_log("Entering get_property on main")
        prop_name = data_dict["property"]
        val = getattr(self, prop_name)
        return {"success": True, "val": val}

    @task_worthy
    def export_data(self, data):
        mdata = self.create_initial_metadata()
        mdata["name"] = "__metadata__"
        full_collection_name = data["full_collection_name"]
        self.db[full_collection_name].insert_one(mdata)
        for docinfo in self.doc_dict.values():
            self.db[full_collection_name].insert_one({"name": docinfo.name,
                                                      "data_rows": docinfo.data_rows,
                                                      "header_list": docinfo.header_list})
        return {"success": True}

    @task_worthy
    def get_doc_dict(self, data):
        return {"success": True, "doc_dict": self.doc_dict}


    @task_worthy
    def get_tile_ids(self, data):
        tile_ids = self.tile_instances.keys()
        return {"success": True, "tile_ids": tile_ids}

    @task_worthy
    def set_property(self, data_dict):
        prop_name = data_dict["property"]
        val = data_dict["val"]
        setattr(self, prop_name, val)
        return

    @task_worthy
    def open_log_window(self, task_data):
        self.console_html = task_data["console_html"]
        if self.project_name is None:
            title = self.short_collection_name + " log"
        else:
            title = self.project_name + " log"
        self.post_task("host", "open_log_window", {"console_html": self.console_html,
                                                   "title": title,
                                                   "main_id": self.my_id})
        return

    # todo download_collection needs rethinking look at how pipe_values handled for a good model
    # todo I'm in the middle of figuring out how to make this work.
    @task_worthy
    def download_collection(self, data):
        new_name = data["new_name"]

        wb = openpyxl.Workbook()
        first = True
        for (doc_name, doc_info) in self.doc_dict.items():
            if first:
                ws = wb.active
                ws.title = doc_name
                first = False
            else:
                ws = wb.create_sheet(title=doc_name)
            data_rows = doc_info.all_sorted_data_rows
            header_list = doc_info.header_list
            for c, header in enumerate(header_list, start=1):
                _ = ws.cell(row=1, column=c, value=header)
            for r, row in enumerate(data_rows, start=2):
                for c, header in enumerate(header_list, start=1):
                    _ = ws.cell(row=r, column=c, value=row[header])
            # noinspection PyUnresolvedReferences
        virtual_notebook = openpyxl.writer.excel.save_virtual_workbook(wb)
        str_io = cStringIO.StringIO()
        str_io.write(virtual_notebook)
        str_io.seek(0)
        encoded_str_io = Binary(cPickle.dumps(str_io))
        self.post_task("host", "send_file_to_client", {"encoded_str_io": encoded_str_io})
        return

    @task_worthy
    def grab_data(self, data):
        doc_name = data["doc_name"]
        return {"doc_name": doc_name,
                "is_shrunk": self.is_shrunk,
                "left_fraction": self.left_fraction,
                "data_rows": self.doc_dict[doc_name].displayed_data_rows,
                "background_colors": self.doc_dict[doc_name].displayed_background_colors,
                "header_list": self.doc_dict[doc_name].header_list,
                "is_last_chunk": self.doc_dict[doc_name].is_last_chunk,
                "is_first_chunk": self.doc_dict[doc_name].is_first_chunk,
                "max_table_size": self.doc_dict[doc_name].max_table_size}

    @task_worthy
    def grab_project_data(self, data_dict):
        doc_name = data_dict["doc_name"]
        return {"doc_name": doc_name,
                "is_shrunk": self.is_shrunk,
                "tile_ids": self.tile_sort_list,
                "left_fraction": self.left_fraction,
                "data_rows": self.doc_dict[doc_name].displayed_data_rows,
                "background_colors": self.doc_dict[doc_name].displayed_background_colors,
                "header_list": self.doc_dict[doc_name].header_list,
                "is_last_chunk": self.doc_dict[doc_name].is_last_chunk,
                "is_first_chunk": self.doc_dict[doc_name].is_first_chunk,
                "max_table_size": self.doc_dict[doc_name].max_table_size,
                "tablespec_dict": self.tablespec_dict(),
                "hidden_columns_list": self.hidden_columns_list,
                "tile_save_results": self.tile_save_results}

    def get_tile_property(self, tile_id, prop_name):
        result = self.post_and_wait(tile_id, 'get_property', {"property": prop_name})["val"]
        return result

    @task_worthy
    def reload_tile(self, ddict):
        self.debug_log("entering reload_tile")
        tile_id = ddict["tile_id"]
        tile_type = self.get_tile_property(tile_id, "tile_type")
        data = {"tile_type": tile_type, "user_id": self.user_id}
        module_code = self.post_and_wait("host", "get_module_code", data)["module_code"]
        list_names = self.post_and_wait("host", "get_list_names", {"user_id": self.user_id})["list_names"]
        reload_dict = copy.copy(self.get_tile_property(tile_id, "current_reload_attrs"))
        saved_options = copy.copy(self.get_tile_property(tile_id, "current_options"))
        reload_dict.update(saved_options)
        result = send_request_to_container(self.tile_instances[tile_id], "load_source",
                                  {"tile_code": module_code, "megaplex_address": self.megaplex_address}).json()
        if not result["success"]:
            raise Exception(result["message_string"])

        form_info = {"current_header_list": self.current_header_list,
                     "pipe_dict": self._pipe_dict,
                     "doc_names": self.doc_names,
                     "list_names": list_names}
        reload_dict["form_info"] = form_info
        reload_dict["main_id"] = self.my_id
        result = send_request_to_container(self.tile_instances[tile_id], "reinstantiate_tile", reload_dict).json()
        if result["success"]:
            return {"success": True, "html": result["form_html"]}
        else:
            raise Exception(result["message_string"])

    @task_worthy
    def grab_chunk_with_row(self, data_dict):
        self.debug_log("Entering grab chunk with row")
        doc_name = data_dict["doc_name"]
        row_id = data_dict["row_id"]
        self.doc_dict[doc_name].move_to_row(row_id)
        return {"doc_name": doc_name,
                "left_fraction": self.left_fraction,
                "is_shrunk": self.is_shrunk,
                "data_rows": self.doc_dict[doc_name].displayed_data_rows,
                "background_colors": self.doc_dict[doc_name].displayed_background_colors,
                "header_list": self.doc_dict[doc_name].header_list,
                "is_last_chunk": self.doc_dict[doc_name].is_last_chunk,
                "is_first_chunk": self.doc_dict[doc_name].is_first_chunk,
                "max_table_size": self.doc_dict[doc_name].max_table_size,
                "actual_row": self.doc_dict[doc_name].get_actual_row(row_id)}

    @task_worthy
    def get_actual_row(self, data):
        doc_name = data["doc_name"]
        row_index = data["row_index"]
        actual_row = self.doc_dict[doc_name].get_actual_row(row_index)
        return {"actual_row": actual_row}

    @task_worthy
    def distribute_events_stub(self, data_dict):
        event_name = data_dict["event_name"]

        # If necessary, have to convert the row_index on the client side to the row_id
        if (data_dict is not None) and ("active_row_index" in data_dict) and ("doc_name" in data_dict):
            if data_dict["active_row_index"] is not None:
                data_dict["active_row_index"] = self.doc_dict[data_dict["doc_name"]].get_id_from_actual_row(
                    data_dict["active_row_index"])
        if "tile_id" in data_dict:
            tile_id = data_dict["tile_id"]
        else:
            tile_id = None
        success = self.distribute_event(event_name, data_dict, tile_id)
        return {"success": success}

    @task_worthy
    def grab_next_chunk(self, data_dict):
        self.debug_log("entering grab next chunk")
        doc_name = data_dict["doc_name"]
        step_amount = self.doc_dict[doc_name].advance_to_next_chunk()
        return {"doc_name": doc_name,
                "data_rows": self.doc_dict[doc_name].displayed_data_rows,
                "background_colors": self.doc_dict[doc_name].displayed_background_colors,
                "header_list": self.doc_dict[doc_name].header_list,
                "is_last_chunk": self.doc_dict[doc_name].is_last_chunk,
                "is_first_chunk": self.doc_dict[doc_name].is_first_chunk,
                "step_size": step_amount}

    @task_worthy
    def grab_previous_chunk(self, data_dict):
        self.debug_log("entering grab previous chunk")
        doc_name = data_dict["doc_name"]
        step_amount = self.doc_dict[doc_name].go_to_previous_chunk()
        return {"doc_name": doc_name,
                "data_rows": self.doc_dict[doc_name].displayed_data_rows,
                "background_colors": self.doc_dict[doc_name].displayed_background_colors,
                "header_list": self.doc_dict[doc_name].header_list,
                "is_last_chunk": self.doc_dict[doc_name].is_last_chunk,
                "is_first_chunk": self.doc_dict[doc_name].is_first_chunk,
                "step_size": step_amount}

    @task_worthy
    def RemoveTile(self, data):
        self._delete_tile_instance(data["tile_id"])
        return None

    @task_worthy
    def CreateColumn(self, data):
        column_name = data["column_name"]
        for doc in self.doc_dict.values():
            doc.header_list.append(column_name)
            for r in doc.data_rows.values():
                r[column_name] = ""
        form_info = {"current_header_list": self.current_header_list,
                     "pipe_dict": self._pipe_dict,
                     "doc_names": self.doc_names,
                     "list_names": self.get_list_names()}
        for tid in self.tile_instances.keys():
            self.post_task(tid, "RebuildTileForms", form_info)
        return None

    @task_worthy
    def SearchTable(self, data):
        self.highlight_table_text(data["text_to_find"])
        return None

    @task_worthy
    def FilterTable(self, data):
        txt = data["text_to_find"]
        self.display_matching_rows_applying_filter(lambda r: self.txt_in_dict(txt, r))
        self.highlight_table_text(txt)
        return None

    @task_worthy
    def DehighlightTable(self, data):
        self.emit_table_message("dehiglightAllCells")
        return None

    @task_worthy
    def UnfilterTable(self, data):
        for doc in self.doc_dict.values():
            doc.current_data_rows = doc.data_rows
            doc.configure_for_current_data()
        self.refill_table()
        return None

    @task_worthy
    def ColorTextInCell(self, data):
        self.emit_table_message("colorTxtInCell", data)
        return None

    @task_worthy
    def SetCellContent(self, data):
        self._set_cell_content(data["doc_name"], data["id"], data["column_header"],
                               data["new_content"], data["cellchange"])
        return None

    @task_worthy
    def TextSelect(self, data):
        self.selected_text = data["selected_text"]
        return None

    @task_worthy
    def SaveTableSpec(self, data):
        new_spec = data["tablespec"]
        self.doc_dict[new_spec["doc_name"]].table_spec = new_spec
        return None

    @task_worthy
    def UpdateSortList(self, data):
        self.tile_sort_list = data["sort_list"]
        return None

    @task_worthy
    def UpdateLeftFraction(self, data):
        self.left_fraction = data["left_fraction"]
        return None

    @task_worthy
    def UpdateTableShrinkState(self, data):
        self.is_shrunk = data["is_shrunk"]
        return None

    @task_worthy
    def PrintToConsole(self, data):
        self.print_to_console(data["message"], True)
        return None

    @task_worthy
    def DisplayCreateErrors(self, data):
        for msg in self.recreate_errors:
            self.debug_log("Got CreateError: " + msg)
            self.print_to_console(msg, True)
        self.recreate_errors = []
        return None

    # called view get_func

    def get_matching_rows(self, filter_function, document_name):
        result = []
        if document_name is not None:
            for r in self.doc_dict[document_name].sorted_data_rows:
                if filter_function(r):
                    result.append(r)
        else:
            for doc in self.doc_dict.keys():
                for r in self.doc_dict[doc].sorted_data_rows:
                    if filter_function(r):
                        result.append(r)
        return result

    @task_worthy
    def display_matching_rows(self, data):
        self.debug_log("Entering display_matching_rows in main.py")
        result = data["result"]
        document_name = data["document_name"]
        if document_name is not None:
            doc = self.doc_dict[document_name]
            doc.current_data_rows = {}
            for (key, val) in doc.data_rows.items():
                if int(key) in result:
                    doc.current_data_rows[key] = val
            doc.configure_for_current_data()
            self.refill_table()
        else:
            for docname, doc in self.doc_dict.items():
                doc.current_data_rows = {}
                self.debug_log("result[docname] is " + str(result[docname]))
                for (key, val) in doc.data_rows.items():
                    if int(key) in result[docname]:
                        doc.current_data_rows[key] = val
                doc.configure_for_current_data()
            self.refill_table()
        return

    @task_worthy
    def update_document(self, data):
        new_data= data["new_data"]
        doc_name = data["document_name"]
        doc = self.doc_dict[doc_name]
        for key, r in new_data.items():
            for c, val in r.items():
                doc.data_rows[key][c] = val
        if doc_name == self.visible_doc_name:
            self.refill_table()
        return {"success": True}

    def display_matching_rows_applying_filter(self, filter_function, document_name=None):
        if document_name is not None:
            doc = self.doc_dict[document_name]
            doc.current_data_rows = {}
            for (key, val) in doc.data_rows.items():
                if filter_function(val):
                    doc.current_data_rows[key] = val
            doc.configure_for_current_data()
            self.refill_table()
        else:
            for docname, doc in self.doc_dict.items():
                doc.current_data_rows = {}
                for (key, val) in doc.data_rows.items():
                    if filter_function(val):
                        doc.current_data_rows[key] = val
                doc.configure_for_current_data()
            self.refill_table()
        return

    def apply_to_rows(self, func, document_name=None):
        if document_name is not None:
            i = 0
            for r in self.doc_dict[document_name].sorted_data_rows:
                new_r = func(r)
                for (key, val) in new_r.items():
                    # self.doc_dict[document_name].data_rows[i][key] = val
                    self._set_cell_content(document_name, i, key, val, cellchange=False)
                i += 1
        else:
            for doc in self.doc_dict.keys():
                i = 0
                for r in self.doc_dict[doc].sorted_data_rows:
                    new_r = func(r)
                    for (key, val) in new_r.itesm():
                        # self.doc_dict[document_name].data_rows[i][key] = val
                        self._set_cell_content(doc, i, key, val, cellchange=False)
                    i += 1
        return

    # _set_cell_content is called from several places
    def _set_cell_content(self, doc_name, the_id, column_header, new_content, cellchange=True):
        doc = self.doc_dict[doc_name]
        the_row = doc.data_rows[str(the_id)]
        old_content = the_row[column_header]
        if new_content != old_content:
            data = {"doc_name": doc_name, "id": the_id, "column_header": column_header,
                    "new_content": new_content, "old_content": old_content}

            # If cellchange is True then we use a CellChange event to handle any updates.
            # Otherwise just change things right here.
            if cellchange:
                self.distribute_event("CellChange", data)
            else:
                self._set_row_column_data(doc_name, the_id, column_header, new_content)
                self._change_list.append(the_id)
            if doc_name == self.visible_doc_name:
                doc = self.doc_dict[doc_name]
                actual_row = doc.get_actual_row(the_id)
                if actual_row is not None:
                    data["row"] = actual_row
                    self.emit_table_message("setCellContent", data)

    def _set_cell_background(self, doc_name, the_id, column_header, color):
        doc = self.doc_dict[doc_name]
        doc.set_background_color(the_id, column_header, color)
        if doc_name == self.visible_doc_name:
            actual_row = doc.get_actual_row(the_id)
            if actual_row is not None:
                data = {"row": actual_row,
                        "column_header": column_header,
                        "color": color}
                self.emit_table_message("setCellBackground", data)

