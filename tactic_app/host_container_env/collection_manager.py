
import re, datetime, sys, os
from collections import OrderedDict
import tempfile
import zipfile
from flask_login import login_required, current_user
from flask import jsonify, render_template, url_for, request, send_file
from users import User
from docker_functions import main_container_info
from tactic_app import app, db, repository_db, use_remote_database
from communication_utils import make_python_object_jsonizable, debinarize_python_object, make_jsonizable_and_compress
from communication_utils import read_temp_data, delete_temp_data
from mongo_accesser import MongoAccessException, NonexistentNameError
import tempfile
# noinspection PyPackageRequirements
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import io
# import cStringIO
import tactic_app
from file_handling import read_csv_file_to_list, read_txt_file_to_list
from file_handling import read_freeform_file, read_excel_file
from users import load_user
from redis_tools import create_ready_block

from js_source_management import js_source_dict, _develop, css_source

from resource_manager import ResourceManager, LibraryResourceManager, repository_user
import loaded_tile_management

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

CHUNK_SIZE = int(int(os.environ.get("CHUNK_SIZE")) / 2)

AUTOSPLIT = False
AUTOSPLIT_SIZE = 10000

import datetime
tstring = datetime.datetime.utcnow().strftime("%Y-%H-%M-%S")


# noinspection PyMethodMayBeStatic,PyBroadException,RegExpRedundantEscape
class CollectionManager(LibraryResourceManager):
    collection_list = "data_collection_names"
    collection_list_with_metadata = "data_collection_names_with_metadata"
    collection_name = "collection_collection_name"
    name_field = "collection_name"

    def add_rules(self):
        app.add_url_rule('/new_notebook', "new_notebook",
                         login_required(self.new_notebook), methods=['get', 'post'])
        app.add_url_rule('/new_notebook_with_data/<temp_data_id>', "new_notebook_with_data",
                         login_required(self.new_notebook_with_data), methods=['get', 'post'])
        app.add_url_rule('/new_notebook_in_context', "new_notebook_in_context",
                         login_required(self.new_notebook_in_context), methods=['get', 'post'])
        app.add_url_rule('/main_collection/<collection_name>', "main",
                         login_required(self.main_collection), methods=['get'])
        app.add_url_rule('/main_collection_in_context', "main_in_context",
                         login_required(self.main_collection_in_context), methods=['get', 'post'])
        app.add_url_rule('/new_project', "new_project",
                         login_required(self.new_project), methods=['get'])
        app.add_url_rule('/new_project_in_context', "new_project_in_context",
                         login_required(self.new_project_in_context), methods=['get', 'post'])
        app.add_url_rule('/create_empty_collection', "create_empty_collection",
                         login_required(self.create_empty_collection), methods=['get', "post"])
        app.add_url_rule('/append_documents_to_collection/<collection_name>/<doc_type>/<library_id>',
                         "append_documents_to_collection",
                         login_required(self.append_documents_to_collection), methods=['get', "post"])
        app.add_url_rule('/delete_resource_list', "delete_resource_list",
                         login_required(self.delete_resource_list), methods=['post'])
        app.add_url_rule('/open_raw/<collection_name>', "open_raw",
                         login_required(self.open_raw), methods=['post', 'get'])
        app.add_url_rule('/duplicate_collection', "duplicate_collection",
                         login_required(self.duplicate_collection), methods=['post', 'get'])
        app.add_url_rule('/download_temp_collection/<download_name>/<temp_id>', "download_temp_collection",
                         login_required(self.download_temp_collection), methods=['post', 'get'])
        app.add_url_rule('/download_collection/<collection_name>/<new_name>', "download_collection",
                         login_required(self.download_collection), methods=['post', 'get'])
        app.add_url_rule('/combine_collections/<base_collection_name>/<collection_to_add>', "combine_collections",
                         login_required(self.combine_collections), methods=['post', 'get'])
        app.add_url_rule('/combine_to_new_collection', "combine_to_new_collection",
                         login_required(self.combine_to_new_collection), methods=['post'])
        app.add_url_rule('/grab_all_list_chunk', "grab_all_list_chunk",
                         login_required(self.grab_all_list_chunk), methods=['get', 'post'])

    def new_notebook_in_context(self):
        user_obj = current_user
        if "temp_data_id" in request.json:
            temp_data_id = request.json["temp_data_id"]
            main_id, rb_id = main_container_info.create_main_container("new_notebook",
                                                                       user_obj.get_id(),
                                                                       user_obj.username)
        else:
            temp_data_id = ""
            main_id, rb_id = main_container_info.create_main_container("new_notebook",
                                                                       user_obj.get_id(),
                                                                       user_obj.username)
        create_ready_block(rb_id, user_obj.username, [main_id, "client"], main_id)
        data_dict = {"success": True,
                     "kind": "notebook-viewer",
                     "res_type": "project",
                     "project_name": "",
                     "resource_name": "new notebook",
                     "ready_block_id": rb_id,
                     "main_id": main_id,
                     "temp_data_id": temp_data_id,
                     "collection_name": "",
                     "doc_names": [],
                     "short_collection_name": "",
                     "doc_type": "notebook",
                     "is_table": False,
                     "is_notebook": True,
                     "is_freeform": False,
                     "is_jupyter": False,
                     "is_project": False,
                     "base_figure_url": url_for("figure_source", tile_id="tile_id", figure_name="X")[:-1]}
        return jsonify(data_dict)

    def new_notebook(self):
        return self.new_notebook_with_data("")

    def new_notebook_with_data(self, temp_data_id):
        return render_template("main_react.html",
                               project_name='',
                               is_new_notebook="True",
                               base_figure_url=url_for("figure_source", tile_id="tile_id", figure_name="X")[:-1],
                               temp_data_id=temp_data_id,
                               develop=str(_develop),
                               theme=current_user.get_theme(),
                               is_jupyter="False",
                               version_string=tstring,
                               css_source=css_source("notebook_app"),
                               module_source=js_source_dict["notebook_app"])

    def main_collection_in_context(self):
        user_obj = current_user
        short_collection_name = request.json["resource_name"]
        main_id, rb_id = main_container_info.create_main_container(short_collection_name, user_obj.get_id(),
                                                                   user_obj.username)
        create_ready_block(rb_id, user_obj.username, [main_id, "client"], main_id)
        doc_dict, doc_mddict, hl_dict, mdata = user_obj.get_all_collection_info(short_collection_name)
        if "_id" in mdata:
            del(mdata["_id"])
        if "type" in mdata and mdata["type"] == "freeform":
            doc_type = "freeform"
        else:
            doc_type = "table"
        doc_names = list(doc_dict.keys())
        data = {
            "success": True,
            "kind": "main-viewer",
            "res_type": "collection",
            "short_collection_name": short_collection_name,
            "resource_name": short_collection_name,
            "collection_name": short_collection_name,
            "main_id": main_id,
            "ready_block_id": rb_id,
            "mdata": mdata,
            "is_project": False,
            "project_name": "",
            "doc_names": doc_names,
            "base_figure_url": url_for("figure_source", tile_id="tile_id", figure_name="X")[:-1],
            "temp_data_id": "",
            "console_html": "",
            "doc_type": doc_type,
            "is_table": doc_type == "table",
            "is_freeform": doc_type == "freeform"
        }
        return jsonify(data)

    def main_collection(self, collection_name):

        return render_template("main_react.html",
                               collection_name=collection_name,
                               window_title=collection_name,
                               project_name="",
                               is_new_notebook="False",
                               theme=current_user.get_theme(),
                               develop=str(_develop),
                               version_string=tstring,
                               css_source=css_source("main_app"),
                               module_source=js_source_dict["main_app"])

    def new_project_in_context(self):
        user_obj = current_user
        main_id, rb_id = main_container_info.create_main_container("", user_obj.get_id(),
                                                                   user_obj.username)
        create_ready_block(rb_id, user_obj.username, [main_id, "client"], main_id)
        doc_type = "none"
        data = {
            "success": True,
            "kind": "main-viewer",
            "res_type": "collection",
            "short_collection_name": "",
            "resource_name": "new project",
            "collection_name": "",
            "main_id": main_id,
            "ready_block_id": rb_id,
            "mdata": "",
            "is_project": False,
            "project_name": "",
            "doc_names": [],
            "base_figure_url": url_for("figure_source", tile_id="tile_id", figure_name="X")[:-1],
            "temp_data_id": "",
            "console_html": "",
            "doc_type": doc_type,
            "is_table": False,
            "is_freeform": False
        }
        return jsonify(data)

    def new_project(self,):

        return render_template("main_react.html",
                               collection_name="",
                               window_title="New Project",
                               project_name="",
                               is_new_notebook="False",
                               theme=current_user.get_theme(),
                               develop=str(_develop),
                               version_string=tstring,
                               css_source=css_source("main_app"),
                               module_source=js_source_dict["main_app"])

    def rename_me(self, old_name):
        try:
            new_name = request.json["new_name"]
            self.db[current_user.collection_collection_name].update_one({"collection_name": old_name},
                                                                        {'$set': {"collection_name": new_name}})
            return jsonify({"success": True, "message": "collection name changed", "alert_type": "alert-success"})
        except Exception as ex:
            return self.get_exception_for_ajax(ex, "Error renaming collection")

    def adjust_ws_col_widths(self, ws, max_col_width):
        def as_text(value):
            if value is None:
                return ""
            return str(value)
        for column_cells in ws.columns:
            wrap = False
            length = max(len(as_text(cell.value)) for cell in column_cells) + 5
            if length > max_col_width:
                length = max_col_width
                wrap = True
            col = ws.column_dimensions[get_column_letter(column_cells[0].column)]
            col.width = length
            if wrap:
                for cell in column_cells:
                    cell.alignment = Alignment(wrap_text=True)
        return

    def open_raw(self, collection_name):
        user_obj = current_user
        coll_dict, doc_mdata_dict, header_list_dict, coll_mdata = user_obj.get_all_collection_info(collection_name,
                                                                                                   return_lists=False)
        doc_type = "freeform" if coll_mdata["type"] == "freeform" else "table"
        if doc_type == "table":
            return "Only Freeform docs can be opened raw"
        return list(coll_dict.values())[0]

    def download_temp_collection(self, download_name, temp_id):
        return self.download_collection("", download_name, temp_id=temp_id)

    # noinspection PyTypeChecker
    def download_collection(self, collection_name, new_name, max_col_width=50, temp_id=None):
        user_obj = current_user
        try:
            coll_dict, doc_mdata_dict, header_list_dict, coll_mdata = user_obj.get_all_collection_info(collection_name,
                                                                                                       return_lists=False,
                                                                                                       temp_id=temp_id)
        except NonexistentNameError:
            return "Collection name not found"
        if temp_id is not None:
            delete_temp_data(self.db, temp_id)

        wb = openpyxl.Workbook()
        first = True
        doc_type = "freeform" if coll_mdata["type"] == "freeform" else "table"

        if doc_type == "freeform":
            if len(coll_dict) ==  1:
                filename, file_extension = os.path.splitext(new_name)
                if len(file_extension) == 0:
                    download_name = new_name + ".txt"
                else:
                    download_name = new_name
                mem = io.BytesIO()
                mem.write(list(coll_dict.values())[0].encode())
                mem.seek(0)
                return send_file(mem,
                                 download_name=download_name,
                                 as_attachment=True)
            else:
                if new_name.endswith(".zip"):
                    download_name = new_name
                else:
                    download_name = new_name + ".zip"
                with tempfile.TemporaryDirectory() as tmpdir:
                    for doc_name, doc_text in coll_dict.items():
                        with open(os.path.join(tmpdir, doc_name + ".txt"), 'w') as file:
                            file.write(doc_text)

                    with zipfile.ZipFile(download_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        for root, dirs, files in os.walk(tmpdir):
                            for file in files:
                                zipf.write(os.path.join(root, file))
                    return send_file(download_name, as_attachment=True)

        if new_name.endswith(".xlsx"):
            download_name = new_name
        else:
            download_name = new_name + ".xlsx"
        for doc_name in coll_dict.keys():
            sheet_name = re.sub(r"[\[\]\*\/\\ \?\:]", r"-", doc_name)[:25]
            if first:
                ws = wb.active
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(title=sheet_name)
            if doc_type == "table":
                data_rows = coll_dict[doc_name]
                header_list = header_list_dict[doc_name]
            else:
                header_list = ["text"]
                data_text = coll_dict[doc_name].splitlines()
                data_rows = {}
                for r, txt in enumerate(data_text):
                    data_rows[str(r)] = {"text": txt}
            for c, header in enumerate(header_list, start=1):
                _ = ws.cell(row=1, column=c, value=header)
                ws.cell(1, c).font = Font(bold=True)
            sorted_int_keys = sorted([int(key) for key in data_rows.keys()])
            for r, _id in enumerate(sorted_int_keys, start=2):
                row = data_rows[str(_id)]
                for c, header in enumerate(header_list, start=1):
                    try:
                        val = re.sub(ILLEGAL_CHARACTERS_RE, " ", str(row[header]))
                    except:
                        val = None
                    _ = ws.cell(row=r, column=c, value=val)
            self.adjust_ws_col_widths(ws, max_col_width)

        tmp = tempfile.NamedTemporaryFile()
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(tmp,
                         download_name=download_name,
                         as_attachment=True)

    def remove_duplicate_collections(self, user_obj=None):
        print("entering remove duplicate collections")
        if user_obj is None:
            user_obj = current_user

        cnames = user_obj.data_collection_names
        already_deleted = []
        for cname in cnames:
            if cnames.count(cname) > 1 and cname not in already_deleted:
                print("removing duplicate collection " + cname)
                user_obj.remove_collection(cname)
                already_deleted.append(cname)

        return {"success": True}

    def db_is_updated(self, is_repo):
        if is_repo:
            user_obj = repository_user
            db_to_use = self.repository_db
        else:
            user_obj = current_user
            db_to_use = self.db

        return user_obj.collection_collection_name in db_to_use.list_collection_names() and \
                db_to_use[user_obj.collection_collection_name].count_documents({}) > 0

    def grab_metadata(self, res_name):
        user_obj = current_user
        return user_obj.get_collection_metadata(res_name)

    def save_metadata(self, res_name, tags, notes):
        current_user.set_collection_metadata(res_name, tags, notes)
        return

    def delete_tag(self, tag):
        cnames_with_metadata = current_user.data_collection_names_with_metadata
        for [res_name, mdata] in cnames_with_metadata:
            if mdata is None:
                continue
            tagstring = mdata["tags"]
            taglist = tagstring.split()
            if tag in taglist:
                taglist.remove(tag)
                mdata["tags"] = " ".join(taglist)
                current_user.set_collection_metadata(res_name, mdata["tags"], mdata["notes"])
        return

    def rename_tag(self, tag_changes):
        cnames_with_metadata = current_user.data_collection_names_with_metadata
        for [res_name, mdata] in cnames_with_metadata:
            tagstring = mdata["tags"]
            taglist = tagstring.split()
            for old_tag, new_tag in tag_changes:
                if old_tag in taglist:
                    taglist.remove(old_tag)
                    if new_tag not in taglist:
                        taglist.append(new_tag)
                    mdata["tags"] = " ".join(taglist)
                    current_user.set_collection_metadata(res_name, mdata["tags"], mdata["notes"])
        return

    def autosplit_doc(self, filename, full_dict):
        sorted_int_keys = sorted([int(key) for key in full_dict.keys()])
        counter = 0
        doc_list = []
        doc_rows = {}
        doc_counter = 1
        for r in sorted_int_keys:
            doc_rows[str(counter)] = full_dict[str(r)]
            counter += 1
            if counter == AUTOSPLIT_SIZE:
                doc_list.append({
                    "name": filename + "__" + str(doc_counter),
                    "data_rows": doc_rows
                })
                counter = 0
                doc_rows = {}
                doc_counter += 1
        if counter > 0:
            doc_list.append({
                "name": filename + "__" + str(doc_counter),
                "data_rows": doc_rows
            })
        return doc_list

    def get_doc_type(self, coll_name):
        user_obj = current_user
        coll_mdata = user_obj.get_collection_metadata(coll_name)
        if "type" in coll_mdata and coll_mdata["type"] == "freeform":
            doc_type = "freeform"
        else:
            doc_type = "table"
        return doc_type

    def combine_collections(self, base_collection_name, collection_to_add):
        try:
            user_obj = current_user

            if base_collection_name not in user_obj.data_collection_names:
                error_string = base_collection_name + " doesn't exist"
                return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})

            if collection_to_add not in user_obj.data_collection_names:
                error_string = base_collection_name + " doesn't exist"
                return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})
            doc_type = self.get_doc_type(base_collection_name)
            coll_dict, dm_dict, hl_dict, coll_mdata = user_obj.get_all_collection_info(collection_to_add)
            if not coll_mdata["type"] == doc_type:
                error_string = "Cannot combine freeform and table collections"
                return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})
            user_obj.append_documents_to_collection(base_collection_name, coll_dict, doc_type, hl_dict, dm_dict)
            user_obj.update_collection_time(base_collection_name)
            return jsonify({"success": True,
                            "message": "Collections successfull combined",
                            "alert_type": "alert-success"})
        except Exception as ex:
            return self.get_traceback_exception_for_ajax(ex, "Error combining collection")

    def create_empty_collection(self):
        collection_name = request.json["collection_name"]
        doc_type = request.json["doc_type"]
        library_id = request.json["library_id"]
        user_obj = current_user

        collection_mdata = loaded_tile_management.create_initial_metadata()
        if "csv_options" in request.json and request.json["csv_options"]:
            collection_mdata["csv_options"] = request.json["csv_options"]
        try:
            result = user_obj.create_empty_collection(collection_name, doc_type, collection_mdata)
            result["title"] = "Collection {} created".format(collection_name)
        except Exception as ex:
            msg = self.extract_short_error_message(ex, "Error creating collection")
            result = {"success": False, "title": "Error creating collection", "content": msg}
            self.send_import_report(result, library_id)
        return result

    def append_documents_to_collection(self, collection_name, doc_type, library_id):
        file_list = []
        for the_file in request.files.values():
            file_list.append(the_file)
        print("** received {} files.".format(len(file_list)))
        if len(file_list) == 0:
            return {"success": "false", "title": "Error creating collection", "content": "No files received"}
        if doc_type == "table":
            result = self.append_table_documents(collection_name, file_list)
        else:
            result = self.append_freeform_documents(collection_name, file_list)
        if result["success"] in ["false", "partial"]:
            self.send_import_report(result, library_id)
        return result

    def append_freeform_documents(self, collection_name, file_list):
        user_obj = current_user
        new_doc_dict = {}
        file_decoding_errors = OrderedDict()
        successful_reads = []
        failed_reads = OrderedDict()

        for the_file in file_list:
            filename, file_extension = os.path.splitext(the_file.filename)
            filename = filename.encode("ascii", "ignore").decode()
            (success, result_txt, encoding, decoding_problems) = read_freeform_file(the_file)
            if not success:  # then result_txt contains an error object
                e = result_txt
                failed_reads[filename] = e["message"]
                continue
            new_doc_dict[filename] = result_txt
            if len(decoding_problems) > 0:
                file_decoding_errors[filename] = decoding_problems

        for dname, doc in new_doc_dict.items():
            try:
                _ = user_obj.append_documents_to_collection(collection_name, {dname: doc}, "freeform")
            except Exception as ex:
                msg = self.get_traceback_message(ex, "Error appending document {}".format(dname))
                failed_reads[dname] = msg
                continue
            successful_reads.append(dname)

        if len(successful_reads) == 0:
            final_success = "false"
        elif len(failed_reads.keys()) > 0:
            final_success = "partial"
        else:
            final_success = "true"

        return {"success": final_success,
                "title": "Collection {} created".format(collection_name),
                "file_decoding_errors": file_decoding_errors,
                "successful_reads": successful_reads,
                "failed_reads": failed_reads}

    def append_table_documents(self, collection_name, file_list):
        user_obj = current_user
        new_doc_dict = {}
        header_list_dict = {}
        file_decoding_errors = OrderedDict()
        successful_reads = []
        failed_reads = OrderedDict()
        known_extensions = [".xlsx", ".csv", ".tsv", ".txt"]
        collection_mdata = user_obj.get_collection_metadata(collection_name)
        if "csv_options" in collection_mdata:
            csv_options = collection_mdata["csv_options"]
        else:
            csv_options = None
        for the_file in file_list:
            filename, file_extension = os.path.splitext(the_file.filename)
            filename = filename.encode("ascii", "ignore").decode()
            if file_extension not in known_extensions:
                failed_reads[filename] = "Invalid file extension " + file_extension
                continue
            decoding_problems = []
            if file_extension == ".xlsx":
                (success, doc_dict, header_dict) = read_excel_file(the_file)

                if not success:  # then doc_dict contains an error object
                    e = doc_dict
                    failed_reads[filename] = e["message"]
                    continue
                new_doc_dict.update(doc_dict)
                header_list_dict.update(header_dict)
            else:
                if file_extension in [".csv", ".tsv"]:
                    (success, row_list, header_list, encoding, decoding_problems) = \
                        read_csv_file_to_list(the_file, csv_options)
                # elif file_extension == ".tsv":
                #     (success, row_list, header_list, encoding, decoding_problems) = read_tsv_file_to_list(the_file)
                elif file_extension == ".txt":
                    (success, row_list, header_list, encoding, decoding_problems) = read_txt_file_to_list(the_file)
                else:
                    failed_reads[filename] = "unkown file extension"
                    continue

                if not success:  # then row_list contains an error object
                    e = row_list
                    failed_reads[filename] = e["message"]
                    continue
                new_doc_dict[filename] = row_list
                header_list_dict[filename] = header_list

            if len(decoding_problems) > 0:
                file_decoding_errors[filename] = decoding_problems
        for dname, doc in new_doc_dict.items():
            try:
                _ = user_obj.append_documents_to_collection(collection_name, {dname: doc}, "table",
                                                            {dname: header_list_dict[dname]})
            except Exception as ex:
                msg = self.extract_short_error_message(ex, "Error appending document {}".format(dname))
                failed_reads[dname] = msg
                continue
            successful_reads.append(dname)

        if len(successful_reads) == 0:
            return {"success": "false",
                    "title": "Failed to read document(s)",
                    "file_decoding_errors": file_decoding_errors,
                    "successful_reads": successful_reads,
                    "failed_reads": failed_reads}

        elif len(failed_reads.keys()) > 0 or len(file_decoding_errors.keys()) > 0:
            final_success = "partial"
            title = "Error(s) reading documents"
        else:
            final_success = "true"
            title = ""

        return {"success": final_success,
                "title": title,
                "file_decoding_errors": file_decoding_errors,
                "successful_reads": successful_reads,
                "failed_reads": failed_reads}

    def delete_resource_list(self):
        try:
            user_obj = current_user
            res_list = request.json["resource_list"]
            for row in res_list:
                res_name = row["name"]
                if row["res_type"] == "collection":
                    user_obj.remove_collection(res_name)
                elif row["res_type"] == "project":
                    user_obj.remove_project(res_name)
                elif row["res_type"] == "tile":
                    self.db[user_obj.tile_collection_name].delete_one({"tile_module_name": res_name})
                elif row["res_type"] == "list":
                    self.db[user_obj.list_collection_name].delete_one({"list_name": res_name})
                elif row["res_type"] == "code":
                    self.db[user_obj.code_collection_name].delete_one({"code_name": res_name})
            return jsonify({"success": True, "message": "Resource(s) successfully deleted",
                            "alert_type": "alert-success"})

        except Exception as ex:
            return self.get_exception_for_ajax(ex, "Error deleting collections")

    def combine_to_new_collection(self):
        try:
            user_obj = current_user
            original_collections = request.json["original_collections"]
            new_name = request.json["new_name"]
            coll_dict, dm_dict, hl_dict, coll_mdata = user_obj.get_all_collection_info(original_collections[0])
            doc_type = coll_mdata["type"]
            if "size" in coll_mdata:
                del coll_mdata["size"]
            user_obj.create_complete_collection(new_name,
                                                coll_dict,
                                                coll_mdata["type"],
                                                dm_dict,
                                                hl_dict,
                                                coll_mdata)
            for col in original_collections[1:]:
                if not self.get_doc_type(col) == doc_type:
                    error_string = "Cannot combine freeform and table collections"
                    return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})

            for col in original_collections[1:]:
                coll_dict, dm_dict, hl_dict, coll_mdata = user_obj.get_all_collection_info(col)
                user_obj.append_documents_to_collection(new_name, coll_dict, doc_type, hl_dict, dm_dict)

            user_obj.update_collection_time(new_name)
            return jsonify({"success": True})

        except Exception as ex:
            return self.get_traceback_exception_for_ajax(ex, "Error combining collection")

    def duplicate_collection(self):
        try:
            user_obj = current_user
            new_res_name = request.json["new_res_name"]
            self.show_um_message("Duplicating collection ...", request.json["library_id"])
            coll_dict, dm_dict, hl_dict, coll_mdata = user_obj.get_all_collection_info(request.json['res_to_copy'])
            if "size" in coll_mdata and coll_mdata["size"] == 0:
                del coll_mdata["size"]
            if "type" in coll_mdata:
                ctype = coll_mdata["type"]
            else:
                ctype = "table"    # For old collections
            result = user_obj.create_complete_collection(new_res_name,
                                                         coll_dict,
                                                         ctype,
                                                         dm_dict,
                                                         hl_dict,
                                                         coll_mdata)
            if not result["success"]:
                result["message"] = result["message"]
                result["alert_type"] = "alert-warning"
                return jsonify(result)
            return jsonify({"success": True})
        except Exception as ex:
            msg = self.get_traceback_message(ex)
            self.add_error_drawer_entry("Error duplicating collection", msg, request.json["library_id"])
            return jsonify({"success": False})

    ### Stuff below here is needed if I mount a Mongo database that hasn't yet
    ### Had data collections updated to the new compact format where they all live in a single collection
    ### This is just what is needed for the minimal thing of running the update
    ### There's also some stuff in mongo_accesser

    def upgrade_user_collections(self, user_obj=None):
        print("*** entering upgrade_user_collections ***")
        if user_obj is None:
            user_obj = current_user
        string_start = user_obj.username + ".data_collection."
        cfilter = {"name": {"$regex": string_start + "(.*)"}}
        old_cnames = self.db.list_collection_names(filter=cfilter)
        failed_conversions = []
        for old_cname in old_cnames:
            try:
                print("converting " + old_cname)
                if old_cname == user_obj.username + ".data_collections":
                    continue
                short_name = re.search(string_start + "(.*)", old_cname).group(1)
                if short_name not in user_obj.data_collection_names:
                    print("converting " + short_name)
                    doc_dict, dm_dict, hl_dict, coll_mdata = user_obj.get_all_collection_info_legacy(short_name)
                    new_save_dict = {"metadata": coll_mdata,
                                     "collection_name": short_name}
                    collection_dict = {"doc_dict": doc_dict,
                                       "doc_mdata_dict": dm_dict,
                                       "header_list_dic": hl_dict}
                    cdict = make_jsonizable_and_compress(collection_dict)
                    new_save_dict["file_id"] = self.fs.put(cdict)
                    self.db[user_obj.collection_collection_name].insert_one(new_save_dict)
                    print("removing " + short_name)
                    user_obj.remove_collection_legacy(short_name)
            except Exception as ex:
                failed_conversions.append(old_cname)
                print(self.get_traceback_message(ex))
        print("failed conversions " + str(failed_conversions))
        return {"success": True}

    # legacy
    def get_collection_size(self, short_cname):
        total = 0
        for doc in self.db[current_user.full_collection_name(short_cname)].find():
            if "file_id" in doc:
                total += self.db["fs.files"].find_one({"_id": doc["file_id"]})["length"]
        return total

    # legacy
    def set_collection_size(self, short_collection_name, size):
        name_exists = short_collection_name in self.data_collections
        if not name_exists:
            return None
        full_collection_name = self.build_data_collection_name(short_collection_name)
        mdata = self.get_collection_metadata_legacy(short_collection_name)
        if mdata is None:
            return
        print("setting collection size for " + full_collection_name)
        self.db[full_collection_name].update_one({"name": "__metadata__"},
                                                 {'$set': {"size": size}})
        return

    # legacy
    def get_collection_size_info(self, short_name, mdata):
        if "size" not in mdata:
            col_size = self.get_collection_size(short_name)
            current_user.set_collection_size(short_name, col_size)
        else:
            col_size = mdata["size"]
        if col_size < 100000:
            size_text = "{}kb".format(round(col_size / 1000, 1))
        else:
            size_text = "{}mb".format(round(col_size / 1000000, 1))
        return col_size, size_text


class RepositoryCollectionManager(CollectionManager):
    rep_string = "repository-"
    is_repository = True

    def grab_metadata(self, res_name):
        user_obj = repository_user
        return user_obj.get_collection_metadata(res_name)

    def add_rules(self):
        pass
