import cPickle
import cStringIO
import datetime
import sys
import copy
import requests

import openpyxl
from bson.binary import Binary
from flask import request, jsonify, render_template, send_file, url_for
from flask_login import current_user, login_required
from flask_socketio import join_room
from tactic_app import app, db, fs, socketio
from tactic_app.shared_dicts import mainwindow_instances, get_tile_code
from tactic_app.shared_dicts import tile_classes, user_tiles, loaded_user_modules
from user_manage_views import project_manager, collection_manager
from tactic_app.docker_functions import send_request_to_container, create_container
from tactic_app.docker_functions import get_address, callbacks, destroy_container
from tactic_app.users import load_user


# The main window should join a room associated with the user
@socketio.on('connect', namespace='/main')
def connected_msg():
    print"client connected"


@socketio.on('join', namespace='/main')
def on_join(data):
    room = data["room"]
    join_room(room)
    print "user joined room " + room

# Views for creating and saving a new project
# As well as for updating an existing project.


@app.route('/save_new_project', methods=['POST'])
@login_required
def save_new_project():
    data_dict = request.json
    data_dict["users_loaded_modules"] = loaded_user_modules[current_user.username]
    data_dict["project_collection_name"] = current_user.project_collection_name
    result = send_request_to_container(data_dict["main_id"], "save_new_project", data_dict)
    project_manager.update_selector_list()
    return jsonify(result.json())


@app.route('/get_list_names', methods=["get", "post"])
def get_list_names():
    user_id = request.json["user_id"]
    the_user = load_user(user_id)
    return jsonify({"list_names": the_user.list_names})


@app.route('/delete_container', methods=["get", "post"])
def delete_container():
    container_id = request.json["container_id"]
    destroy_container(container_id)
    return jsonify({"success": True})


@app.route('/get_list', methods=["get", "post"])
def get_list():
    user_id = request.json["user_id"]
    list_name = request.json["list_name"]
    the_user = load_user(user_id)
    return jsonify({"the_list": the_user.get_list(list_name)})


@app.route('/set_visible_doc/<main_id>/<doc_name>', methods=['get', 'post'])
@login_required
def set_visible_doc(main_id, doc_name):
    data_dict = {}
    data_dict["doc_name"] = doc_name
    send_request_to_container(main_id, "set_visible_doc", data_dict)
    return jsonify({"success": True})


@app.route("/add_blank_console_text/<main_id>", methods=["GET"])
@login_required
def add_blank_console_text(main_id):
    try:
        print_string = "<div contenteditable='true'></div>"
        data_dict = {"print_string": print_string}
        send_request_to_container(main_id, "print_to_console", data_dict)
        return jsonify({"success": True})
    except:
        return jsonify({"success": False, "message": "Error creating console text area"})


def set_mainwindow_property(main_id, prop_name, prop_value):
    result = send_request_to_container(main_id, "set_property", {"property": prop_name, "val": prop_value}).json()
    return result


def get_mainwindow_property(main_id, prop_name):
    result = send_request_to_container(main_id, "get_property", {"property": prop_name}).json()
    return result["val"]


@app.route("/send_log_html/<main_id>", methods=["POST"])
@login_required
def send_log_html(main_id):
    try:
        console_html = request.json["console_html"]
        set_mainwindow_property(main_id, "console_html", console_html)
        return jsonify({"success": True})
    except:
        return jsonify({"success": False, "message": "Error opening log wndow"})


@app.route("/open_log_window/<main_id>", methods=["GET", "POST"])
@login_required
def open_log_window(main_id):
    if get_mainwindow_property(main_id, "project_name") is None:
        title = get_mainwindow_property(main_id, "short_collection_name") + " log"
    else:
        title = get_mainwindow_property(main_id, "project_name") + " log"
    console_html = get_mainwindow_property(main_id, "console_html").decode("utf-8", "ignore").encode("ascii")
    return render_template("log_window_template.html", window_title=title, console_html=console_html)


@app.route('/update_project', methods=['POST'])
@login_required
def update_project():
    data_dict = request.json
    data_dict["users_loaded_modules"] = loaded_user_modules[current_user.username]
    data_dict["project_collection_name"] = current_user.project_collection_name
    result = send_request_to_container(data_dict["main_id"], "update_project", data_dict)
    project_manager.update_selector_list()
    return jsonify(result.json())


# todo export_table doesn't work because it's trying to send doc_dict, which isn't serializable
# todo probably give /get_property ability to pickle and convert to binary
@app.route('/export_data', methods=['POST'])
@login_required
def export_data():
    data_dict = request.json
    doc_dict = get_mainwindow_property(data_dict["main_id"], "doc_dict")
    full_collection_name = current_user.build_data_collection_name(data_dict['export_name'])
    for docinfo in doc_dict.values():
        db[full_collection_name].insert_one({"name": docinfo.name,
                                             "data_rows": docinfo.data_rows,
                                             "header_list": docinfo.header_list})
    collection_manager.update_selector_list()
    return jsonify({"success": True, "message": "Data Successfully Exported"})


# todo download_table needs rewriting
@app.route('/download_table/<main_id>/<new_name>', methods=['GET', 'POST'])
@login_required
def download_table(main_id, new_name):
    mw = mainwindow_instances[main_id]
    doc_info = mw.doc_dict[mw.visible_doc_name]
    data_rows = doc_info.all_sorted_data_rows
    header_list = doc_info.header_list
    str_io = cStringIO.StringIO()
    for header in header_list[:-1]:
        str_io.write(str(header) + ',')
    str_io.write(str(header_list[-1]) + '\n')
    for row in data_rows:
        for header in header_list[:-1]:
            str_io.write(str(row[header]) + ",")
        str_io.write(str(row[header_list[-1]]) + '\n')
    str_io.seek(0)
    return send_file(str_io,
                     attachment_filename=new_name,
                     as_attachment=True)


# todo download collection needs rewriting
@app.route('/download_collection/<main_id>/<new_name>', methods=['GET', 'POST'])
@login_required
def download_collection(main_id, new_name):
    mw = mainwindow_instances[main_id]
    wb = openpyxl.Workbook()
    first = True
    for (doc_name, doc_info) in mw.doc_dict.items():
        if first:
            ws = wb.active
            ws.title = doc_name
            first = False
        else:
            ws = wb.create_sheet(title=doc_name)
        data_rows = doc_info.all_sorted_data_rows
        header_list = doc_info.header_list
        for c, header in enumerate(header_list, start=1):
            _ = ws.cell(row=1, column=c, value=header)
        for r, row in enumerate(data_rows, start=2):
            for c, header in enumerate(header_list, start=1):
                _ = ws.cell(row=r, column=c, value=row[header])
    # noinspection PyUnresolvedReferences
    virtual_notebook = openpyxl.writer.excel.save_virtual_workbook(wb)
    str_io = cStringIO.StringIO()
    str_io.write(virtual_notebook)
    str_io.seek(0)
    return send_file(str_io,
                     attachment_filename=new_name,
                     as_attachment=True)

# Views for reading data from the database and
# passing back to the client.


@app.route('/grab_data/<main_id>/<doc_name>', methods=['get', 'post'])
@login_required
def grab_data(main_id, doc_name):
    data_dict = {"doc_name": doc_name}
    result = send_request_to_container(main_id, "grab_data", data_dict)
    return jsonify(result.json())


@app.route('/grab_chunk_with_row', methods=['get', 'post'])
@login_required
def grab_chunk_with_row():
    data_dict = request.json
    main_id = data_dict["main_id"]
    result = send_request_to_container(main_id, "grab_chunk_with_row", data_dict)
    return jsonify(result.json())


@app.route('/grab_next_chunk/<main_id>/<doc_name>', methods=['get'])
@login_required
def grab_next_chunk(main_id, doc_name):
    data_dict = {"doc_name": doc_name}
    result = send_request_to_container(main_id, "grab_next_chunk", data_dict)
    return jsonify(result.json())


@app.route('/grab_previous_chunk/<main_id>/<doc_name>', methods=['get'])
@login_required
def grab_previous_chunk(main_id, doc_name):
    data_dict = {"doc_name": doc_name}
    result = send_request_to_container(main_id, "grab_previous_chunk", data_dict)
    return jsonify(result.json())


@app.route('/grab_project_data/<main_id>/<doc_name>', methods=['get'])
@login_required
def grab_project_data(main_id, doc_name):
    data_dict = {"doc_name": doc_name}
    result = send_request_to_container(main_id, "grab_project_data", data_dict)
    return jsonify(result.json())


@app.route('/get_menu_template', methods=['get'])
@login_required
def get_menu_template():
    return send_file("templates/menu_template.html")


@app.route('/get_table_templates', methods=['get'])
@login_required
def get_table_templates():
    return send_file("templates/table_templates.html")


@app.route('/remove_mainwindow/<main_id>', methods=['get', 'post'])
@login_required
def remove_mainwindow(main_id):
    data_dict = {"main_id": main_id}
    response = send_request_to_container(main_id, "get_tile_ids", data_dict)
    tile_ids = response.json()["tile_ids"]
    for tile_id in tile_ids:
        destroy_container(tile_id)
    destroy_container(main_id)
    return jsonify({"success": True})


@app.route('/get_tile_types', methods=['GET'])
@login_required
def get_tile_types():
    tile_types = {}
    for (category, the_dict) in tile_classes.items():
        tile_types[category] = the_dict.keys()

    if current_user.username in user_tiles:
        for (category, the_dict) in user_tiles[current_user.username].items():
            if category not in tile_types:
                tile_types[category] = []
            tile_types[category] += the_dict.keys()
    result = {"tile_types": tile_types}
    return jsonify(result)


@app.route('/distribute_events/<event_name>', methods=['get', 'post'])
@login_required
def distribute_events_stub(event_name):
    data_dict = request.json
    main_id = request.json["main_id"]
    response = send_request_to_container(main_id, "distribute_events/" + event_name, data_dict)
    return jsonify({"response": response.json()})


@app.route("/request_collection", methods=['GET', 'POST'])
@login_required
def request_collection():
    the_collection = db[request.json["collection_name"]]
    return jsonify({"the_collection": the_collection})


@app.route("/emit_table_message", methods=['GET', 'POST'])
def emit_table_message():
    print "entering emit_table_message on the host"
    data = copy.copy(request.json)
    print "message is " + str(data)
    socketio.emit("table-message", data, namespace='/main', room=data["main_id"])
    return jsonify({"success": True})


@app.route("/emit_tile_message", methods=['GET', 'POST'])
def emit_tile_message():
    print "entering emit_table_message on the host"
    data = copy.copy(request.json)
    print "message is " + str(data)
    socketio.emit("tile-message", data, namespace='/main', room=data["main_id"])
    return jsonify({"success": True})


@app.route("/socketio_emit/<msg>", methods=['GET', 'POST'])
def socketio_emit(msg):
    data = copy.copy(request.json)
    socketio.emit(msg, data, namespace='/main', room=data["main_id"])
    return jsonify({"success": True})


@app.route("/handle_callback", methods=['GET', 'POST'])
def handle_callback():
    print "entering handle_callback on the host"
    data = copy.copy(request.json)
    if "host_callback_id" in data:
        data = callbacks[data["host_callback_id"]](data)
        del callbacks[data["host_callback_id"]]
    if "jcallback_id" in data:
        socketio.emit("handle-callback", data, namespace='/main', room=data["main_id"])
    return jsonify({"success": True})


@app.route('/figure_source/<tile_id>/<figure_name>', methods=['GET', 'POST'])
@login_required
def figure_source(tile_id, figure_name):

    encoded_img = send_request_to_container(tile_id, "get_image/" + figure_name, {}).json()["img"]
    img = cPickle.loads(encoded_img.decode("utf-8", "ignore").encode("ascii"))
    img_file = cStringIO.StringIO()
    img_file.write(img)
    img_file.seek(0)
    return send_file(img_file, mimetype='image/png')


# todo deal with data_source, part of base_data_url, create_data_source
@app.route('/data_source/<main_id>/<tile_id>/<data_name>', methods=['GET'])
@login_required
def data_source(main_id, tile_id, data_name):
    try:
        the_data = mainwindow_instances[main_id].tile_instances[tile_id].data_dict[data_name]
        return jsonify({"success": True, "data": the_data})
    except:
        error_string = str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        mainwindow_instances[main_id].handle_exception("Error getting data " + error_string)
        return jsonify({"success": False})


# noinspection PyUnresolvedReferences
@app.route('/create_tile_request', methods=['GET', 'POST'])
@login_required
def create_tile_request():
    def create_tile_callback(ldata_dict):
        tile_id = ldata_dict["tile_id"]
        form_html = ldata_dict["form_html"]
        tname = data_dict["tile_name"]
        the_html = render_template("tile.html", tile_id=tile_id,
                                   tile_name=tname,
                                   form_text=form_html)
        ddict = copy.copy(ldata_dict)
        ddict["success"] = True
        ddict["html"] = the_html
        ddict["tile_id"] = tile_id
        return ddict
    data_dict = request.json
    main_id = data_dict["main_id"]
    tile_type = data_dict["tile_type"]
    # noinspection PyBroadException
    try:
        tile_container_id = create_container("tactic_tile_image", network_mode="bridge")["Id"]
        module_code = get_tile_code(tile_type)
        send_request_to_container(tile_container_id, "load_source", {"tile_code": module_code})
        tile_container_address = get_address(tile_container_id, "bridge")

        data_dict["tile_id"] = tile_container_id
        data_dict["tile_container_address"] = tile_container_address
        send_request_to_container(main_id, "create_tile_instance",
                                  data_dict, callback=create_tile_callback)
        return jsonify({"success": True})

    except:
        error_string = str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        # mainwindow_instances[main_id].handle_exception("Error creating tile " + error_string)
        return jsonify({"success": False})


# todo work on reload_tiles
@app.route('/reload_tile/<tile_id>', methods=['GET', 'POST'])
@login_required
def reload_tile(tile_id):
    save_attrs_for_reload = ["tile_id", "tile_name", "header_height", "front_height", "front_width", "back_height", "back_width",
                  "tda_width", "tda_height", "width", "height",
                  "full_tile_width", "full_tile_height", "is_shrunk", "configured"]
    try:
        main_id = request.json["main_id"]
        mw = mainwindow_instances[main_id]
        old_instance = mw.tile_instances[tile_id]
        old_instance.kill()
        saved_options = {}
        for option in old_instance.options:
            attr = option["name"]
            if hasattr(old_instance, attr):
                saved_options[attr] = getattr(old_instance, attr)
        tile_type = old_instance.tile_type
        tile_name = old_instance.tile_tname
        new_cls = get_tile_class(current_user.username, tile_type)
        new_instance = new_cls(main_id, tile_id, tile_name)
        for attr, val in saved_options.items():
            setattr(new_instance, attr, val)
        for attr in save_attrs_for_reload:
            setattr(new_instance, attr, getattr(old_instance, attr))
        form_html = new_instance.create_form_html()
        mw.tile_instances[tile_id] = new_instance
        new_instance.start()
        return jsonify({"success": True, "html": form_html})
    except:
        error_string = str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        mainwindow_instances[main_id].handle_exception("Error reloading tile " + error_string)
        return jsonify({"success": False})


@app.route('/create_tile_from_save_request/<tile_id>', methods=['GET', 'POST'])
@login_required
def create_tile_from_save_request(tile_id):
    main_id = request.json["main_id"]
    tile_save_result = send_request_to_container(main_id, "get_saved_tile_info/" + tile_id, {}).json()
    tile_save_result["tile_id"] = tile_id # a little silly that I need to do that but requies less changes this way
    return jsonify(tile_save_result)

