
import re, os
import zipfile
from flask_login import login_required, current_user
from flask import render_template, url_for, request, send_file
from tactic_app import app
from mongo_accesser import NonexistentNameError
import tempfile
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import io
from tactic_logging import log
from utils import utcnow

from js_source_management import js_source_dict, _develop, css_source

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

AUTOSPLIT = False
AUTOSPLIT_SIZE = 10000

tstring = utcnow().strftime("%Y-%H-%M-%S")

@app.route('/new_notebook', methods=['get'])
@login_required
def new_notebook():
    return render_template("main_react.html",
                           project_name='',
                           is_new_notebook="True",
                           read_only="False",
                           is_repository="False",
                           base_figure_url=url_for("figure_source", tile_id="tile_id", figure_name="X")[:-1],
                           temp_data_id="",
                           develop=str(_develop),
                           has_openapi_key=current_user.has_openapi_key,
                           is_jupyter="False",
                           version_string=tstring,
                           css_source=css_source("notebook_app"),
                           module_source=js_source_dict["notebook_app"])


@app.route('/new_notebook_with_data/<temp_data_id>', methods=['get'])
@login_required
def new_notebook_with_data(temp_data_id):
    return render_template("main_react.html",
                           project_name='',
                           is_new_notebook="True",
                           read_only="False",
                           is_repository="False",
                           base_figure_url=url_for("figure_source", tile_id="tile_id", figure_name="X")[:-1],
                           temp_data_id=temp_data_id,
                           develop=str(_develop),
                           has_openapi_key=current_user.has_openapi_key,
                           is_jupyter="False",
                           version_string=tstring,
                           css_source=css_source("notebook_app"),
                           module_source=js_source_dict["notebook_app"])

@app.route('/main_collection/<collection_name>', methods=['get'])
@login_required
def main_collection(collection_name):
    return render_template("main_react.html",
                           collection_name=collection_name,
                           window_title=collection_name,
                           project_name="",
                           read_only="False",
                           is_repository="False",
                           is_new_notebook="False",
                           develop=str(_develop),
                           has_openapi_key=current_user.has_openapi_key,
                           version_string=tstring,
                           css_source=css_source("main_app"),
                           module_source=js_source_dict["main_app"])

@app.route('/new_project', methods=['get'])
@login_required
def new_project():
    return render_template("main_react.html",
                           collection_name="",
                           window_title="New Project",
                           project_name="",
                           is_new_notebook="False",
                           read_only="False",
                           is_repository="False",
                           has_openapi_key=current_user.has_openapi_key,
                           develop=str(_develop),
                           version_string=tstring,
                           css_source=css_source("main_app"),
                           module_source=js_source_dict["main_app"])


@app.route('/append_documents_to_collection/<collection_name>/<doc_type>/<library_id>', methods=['get', 'post'])
def append_documents_to_collection(collection_name, doc_type, library_id):
    user_obj = current_user
    file_list = []
    for the_file in request.files.values():
        file_list.append(the_file)
    log.debug("received files for appending to collection", collection_name=collection_name, num_files=len(file_list))
    if len(file_list) == 0:
        return {"success": "false", "title": "Error creating collection", "content": "No files received"}
    if doc_type == "table":
        result = user_obj.append_table_documents(collection_name, file_list)
    else:
        result = user_obj.append_freeform_documents(collection_name, file_list)
    if result["success"] in ["false", "partial"]:
        user_obj.send_import_report(result, library_id)
    return result


@app.route('/download_temp_collection/<download_name>/<temp_id>', methods=['post', 'get'])
def download_temp_collection(download_name, temp_id):
    return download_collection("", download_name, temp_id=temp_id)


# def delete_temp_data(db, unique_id, fs=None):
#     save_dict = read_temp_data(db, unique_id)
#     db["temp_data"].delete_one({"unique_id": unique_id})
#     if fs is not None and "file_id" in save_dict:
#         fs.delete(save_dict["file_id"])
#     return

@app.route('/download_collection/<collection_name>/<new_name>', methods=['post', 'get'])
def download_collection(collection_name, new_name, max_col_width=50, temp_id=None):
    user_obj = current_user
    try:
        coll_dict, doc_mdata_dict, header_list_dict, coll_mdata = user_obj.get_all_collection_info(collection_name,
                                                                                                   return_lists=False,
                                                                                                   temp_id=temp_id)
    except NonexistentNameError:
        log.exception("Collection name not found: " + collection_name)
        return "Collection name not found"

    if temp_id is not None:
        user_obj.delete_temp_data(temp_id)

    wb = openpyxl.Workbook()
    first = True
    doc_type = "freeform" if coll_mdata["type"] == "freeform" else "table"

    if doc_type == "freeform":
        if len(coll_dict) ==  1:
            filename, file_extension = os.path.splitext(new_name)
            if len(file_extension) == 0:
                download_name = new_name + ".txt"
            else:
                download_name = new_name
            mem = io.BytesIO()
            mem.write(list(coll_dict.values())[0].encode())
            mem.seek(0)
            return send_file(mem,
                             download_name=download_name,
                             as_attachment=True)
        else:
            if new_name.endswith(".zip"):
                download_name = new_name
            else:
                download_name = new_name + ".zip"
            with tempfile.TemporaryDirectory() as tmpdir:
                for doc_name, doc_text in coll_dict.items():
                    with open(os.path.join(tmpdir, doc_name + ".txt"), 'w') as file:
                        file.write(doc_text)

                with zipfile.ZipFile(download_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for root, dirs, files in os.walk(tmpdir):
                        for file in files:
                            zipf.write(os.path.join(root, file))
                return send_file(download_name, as_attachment=True)

    if new_name.endswith(".xlsx"):
        download_name = new_name
    else:
        download_name = new_name + ".xlsx"
    for doc_name in coll_dict.keys():
        sheet_name = re.sub(r"[\[\]\*\/\\ \?\:]", r"-", doc_name)[:25]
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        if doc_type == "table":
            data_rows = coll_dict[doc_name]
            header_list = header_list_dict[doc_name]
        else:
            header_list = ["text"]
            data_text = coll_dict[doc_name].splitlines()
            data_rows = {}
            for r, txt in enumerate(data_text):
                data_rows[str(r)] = {"text": txt}
        for c, header in enumerate(header_list, start=1):
            _ = ws.cell(row=1, column=c, value=header)
            ws.cell(1, c).font = Font(bold=True)
        sorted_int_keys = sorted([int(key) for key in data_rows.keys()])
        for r, _id in enumerate(sorted_int_keys, start=2):
            row = data_rows[str(_id)]
            for c, header in enumerate(header_list, start=1):
                try:
                    val = re.sub(ILLEGAL_CHARACTERS_RE, " ", str(row[header]))
                except:
                    val = None
                _ = ws.cell(row=r, column=c, value=val)
        adjust_ws_col_widths(ws, max_col_width)

    tmp = tempfile.NamedTemporaryFile()
    wb.save(tmp.name)
    tmp.seek(0)
    return send_file(tmp,
                     download_name=download_name,
                     as_attachment=True)

def adjust_ws_col_widths(ws, max_col_width):
    def as_text(value):
        if value is None:
            return ""
        return str(value)
    for column_cells in ws.columns:
        wrap = False
        length = max(len(as_text(cell.value)) for cell in column_cells) + 5
        if length > max_col_width:
            length = max_col_width
            wrap = True
        col = ws.column_dimensions[get_column_letter(column_cells[0].column)]
        col.width = length
        if wrap:
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True)
    return

def remove_duplicate_collections(user_obj=None):
    if user_obj is None:
        user_obj = current_user

    cnames = user_obj.data_collection_names
    already_deleted = []
    for cname in cnames:
        if cnames.count(cname) > 1 and cname not in already_deleted:
            log.debug("removing duplicate collection", collection_name=cname)
            user_obj.remove_collection(cname)
            already_deleted.append(cname)

    return {"success": True}


### Not sure what this db_is_update is for

# def db_is_updated(self, is_repo):
#     if is_repo:
#         user_obj = repository_user
#         db_to_use = self.repository_db
#     else:
#         user_obj = current_user
#         db_to_use = self.db
#
#     return user_obj.collection_collection_name in db_to_use.list_collection_names() and \
#             db_to_use[user_obj.collection_collection_name].count_documents({}) > 0

    ### Stuff below here is needed if I mount a Mongo database that hasn't yet
    ### Had data collections updated to the new compact format where they all live in a single collection
    ### This is just what is needed for the minimal thing of running the update
    ### There's also some stuff in mongo_accesser
    #
    # def upgrade_user_collections(self, user_obj=None):
    #     print("*** entering upgrade_user_collections ***")
    #     if user_obj is None:
    #         user_obj = current_user
    #     string_start = user_obj.username + ".data_collection."
    #     cfilter = {"name": {"$regex": string_start + "(.*)"}}
    #     old_cnames = self.db.list_collection_names(filter=cfilter)
    #     failed_conversions = []
    #     for old_cname in old_cnames:
    #         try:
    #             print("converting " + old_cname)
    #             if old_cname == user_obj.username + ".data_collections":
    #                 continue
    #             short_name = re.search(string_start + "(.*)", old_cname).group(1)
    #             if short_name not in user_obj.data_collection_names:
    #                 print("converting " + short_name)
    #                 doc_dict, dm_dict, hl_dict, coll_mdata = user_obj.get_all_collection_info_legacy(short_name)
    #                 new_save_dict = {"metadata": coll_mdata,
    #                                  "collection_name": short_name}
    #                 collection_dict = {"doc_dict": doc_dict,
    #                                    "doc_mdata_dict": dm_dict,
    #                                    "header_list_dic": hl_dict}
    #                 cdict = make_jsonizable_and_compress(collection_dict)
    #                 new_save_dict["file_id"] = self.fs.put(cdict)
    #                 self.db[user_obj.collection_collection_name].insert_one(new_save_dict)
    #                 print("removing " + short_name)
    #                 user_obj.remove_collection_legacy(short_name)
    #         except Exception as ex:
    #             failed_conversions.append(old_cname)
    #             print(self.get_traceback_message(ex))
    #     print("failed conversions " + str(failed_conversions))
    #     return {"success": True}