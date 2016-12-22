import datetime
import sys
import re

import os
from flask import render_template, request, jsonify, send_file, url_for
from flask_login import login_required, current_user
from flask_socketio import join_room
import openpyxl
import cStringIO

import tactic_app
from tactic_app import app, db, fs, socketio, use_ssl, mongo_uri  # global_stuff
from tactic_app.file_handling import read_csv_file_to_dict, read_tsv_file_to_dict, read_txt_file_to_dict, load_a_list
from tactic_app.file_handling import read_freeform_file

from tactic_app.global_tile_management import global_tile_manager
from tactic_app.users import User, copy_between_accounts
from tactic_app.docker_functions import send_direct_request_to_container

from tactic_app.docker_functions import create_container, get_address
from tactic_app.integrated_docs import api_html, api_dict_by_category, ordered_api_categories
from tactic_app.tile_code_parser import get_functions_full_code
import traceback

AUTOSPLIT = True
AUTOSPLIT_SIZE = 10000


def start_spinner(user_id=None):
    if user_id is None:
        user_id = current_user.get_id()
    socketio.emit('start-spinner', {}, namespace='/user_manage', room=user_id)


def stop_spinner(user_id=None):
    if user_id is None:
        user_id = current_user.get_id()
    socketio.emit('stop-spinner', {}, namespace='/user_manage', room=user_id)


def doflash(message, alert_type='alert-info', user_id=None):
    if user_id is None:
        user_id = current_user.get_id()
    data = {"message": message, "alert_type": alert_type}
    socketio.emit('stop-spinner', data, namespace='/user_manage', room=user_id)


@app.route('/get_resource_names/<res_type>', methods=['get', 'post'])
@login_required
def get_resource_names(res_type):
    manager = managers[res_type][0]
    resource_names = manager.get_resource_list()
    return jsonify({"success": True, "resource_names": resource_names})


@app.route('/get_repository_resource_names/<res_type>', methods=['get', 'post'])
@login_required
def get_repository_resource_names(res_type):
    manager = managers[res_type][1]
    resource_names = manager.get_resource_list()
    return jsonify({"success": True, "resource_names": resource_names})


# noinspection PyMethodMayBeStatic
class ResourceManager(object):
    is_repository = False
    rep_string = ""
    collection_list = ""
    collection_list_with_metadata = ""
    collection_name = ""
    name_field = ""

    def __init__(self, res_type):
        self.res_type = res_type
        self.add_rules()

    def handle_exception(self, ex, special_string=None):
        if special_string is None:
            template = "An exception of type {0} occured. Arguments:\n{1!r}\n"
        else:
            template = special_string + "\n" + "An exception of type {0} occurred. Arguments:\n{1!r}\n"
        error_string = template.format(type(ex).__name__, ex.args)
        error_string += traceback.format_exc()
        doflash("error_string", alert_type="alert-warning")
        return

    def add_rules(self):
        print "not implemented"

    def update_selector_list(self, select=None, user_obj=None):
        if user_obj is None:
            user_obj = current_user

        if self.is_repository:
            socketio.emit('update-selector-list',
                          {"html": self.request_update_selector_list(user_obj=repository_user),
                           "res_type": "repository-" + self.res_type},
                          namespace='/user_manage', room=user_obj.get_id())
        elif select is None:
            socketio.emit('update-selector-list',
                          {"html": self.request_update_selector_list(user_obj=user_obj), "res_type": self.res_type},
                          namespace='/user_manage', room=user_obj.get_id())
        else:
            socketio.emit('update-selector-list',
                          {"html": self.request_update_selector_list(user_obj=user_obj),
                           "select": select,
                           "res_type": self.res_type},
                          namespace='/user_manage', room=user_obj.get_id())

    def get_resource_list(self):
        if self.is_repository:
            user_obj = repository_user
        else:
            user_obj = current_user
        return getattr(user_obj, self.collection_list)

    def get_resource_list_with_metadata(self, user_obj=None):
        if user_obj is None:
            if self.is_repository:
                user_obj = repository_user
            else:
                user_obj = current_user
        return getattr(user_obj, self.collection_list_with_metadata)

    def request_update_selector_list(self, user_obj=None):
        res_list_with_metadata = self.get_resource_list_with_metadata(user_obj)
        res_array = self.build_resource_array(res_list_with_metadata)
        result = self.build_html_table_from_data_list(res_array)
        return result

    # I don't think the timeout variable has any effect.
    def show_um_message(self, message, user_manage_id, timeout=3):
        data = {"message": message, "timeout": timeout}
        socketio.emit('show-status-msg', data, namespace='/user_manage', room=user_manage_id)

    def clear_um_message(self, user_manage_id):
        socketio.emit('clear-status-msg', {}, namespace='/user_manage', room=user_manage_id)

    def build_html_table_from_data_list(self, data_list, title=None):
        the_html = "<table class='tile-table table sortable table-striped table-bordered table-condensed'>"
        if title is not None:
            the_html += "<caption>{0}</caption>".format(title)
        the_html += "<thead><tr>"
        for c in data_list[0]:
            the_html += "<th>{0}</th>".format(c)
        the_html += "</tr><tbody>"
        for r in data_list[1:]:
            the_html += "<tr class='selector-button {0}-selector-button' id='{0}-selector-{1}'>".format(self.res_type,
                                                                                                        r[0])
            for c in r:
                if isinstance(c, list):
                    the_html += "<td sorttable_customkey='{0}'>{1}</td>".format(c[1], c[0])
                else:
                    the_html += "<td>{0}</td>".format(c)
            the_html += "</tr>"

        the_html += "</tbody></table>"
        return the_html

    def build_resource_array(self, res_list):
        larray = [["Name", "Created", "Updated", "Tags"]]
        for res_item in res_list:
            mdata = res_item[1]
            if mdata is None:
                datestring = ""
                tagstring = ""
                updatestring = ""
                datestring_for_sort = ""
                updatestring_for_sort = ""
            else:
                if "datetime" in mdata:
                    datestring = mdata["datetime"].strftime("%b %d, %Y, %H:%M")
                    datestring_for_sort = mdata["datetime"].strftime("%Y%m%d%H%M%S")
                else:
                    datestring = ""
                    datestring_for_sort = ""
                if "updated" in mdata:
                    updatestring = mdata["updated"].strftime("%b %d, %Y, %H:%M")
                    updatestring_for_sort = mdata["updated"].strftime("%Y%m%d%H%M%S")
                else:
                    updatestring = datestring
                    updatestring_for_sort = datestring_for_sort
                tagstring = str(mdata["tags"])
            larray.append([res_item[0], [datestring, datestring_for_sort],
                           [updatestring, updatestring_for_sort], tagstring])
        return larray


def get_manager_for_type(res_type, is_repository=False):

    if is_repository:
        return managers[res_type][1]
    else:
        return managers[res_type][0]


# noinspection PyBroadException
@app.route('/copy_from_repository', methods=['GET', 'POST'])
@login_required
def copy_from_repository():
    res_type = request.json["res_type"]
    new_res_name = request.json['new_res_name']
    res_name = request.json['res_name']
    result = copy_between_accounts(repository_user, current_user, res_type, new_res_name, res_name)
    manager = get_manager_for_type(res_type)
    manager.update_selector_list(select=new_res_name)
    return result


# noinspection PyBroadException
@app.route('/send_to_repository', methods=['GET', 'POST'])
@login_required
def send_to_repository():
    res_type = request.json['res_type']
    new_res_name = request.json['new_res_name']
    res_name = request.json['res_name']
    result = copy_between_accounts(current_user, repository_user, res_type, new_res_name, res_name)
    manager = get_manager_for_type(res_type, is_repository=True)
    manager.update_selector_list(select=new_res_name)
    return result


@app.route('/request_update_selector_list/<res_type>', methods=['GET'])
@login_required
def request_update_selector_list(res_type):
    return managers[res_type][0].request_update_selector_list()


@app.route('/request_update_repository_selector_list/<res_type>', methods=['GET'])
@login_required
def request_update_repository_selector_list(res_type):
    return managers[res_type][1].request_update_selector_list()


# noinspection PyMethodMayBeStatic
class ListManager(ResourceManager):
    collection_list = "list_names"
    collection_list_with_metadata = "list_names_with_metadata"
    collection_name = "list_collection_name"
    name_field = "list_name"

    def add_rules(self):
        app.add_url_rule('/view_list/<list_name>', "view_list", login_required(self.view_list), methods=['get'])
        app.add_url_rule('/repository_view_list/<list_name>', "repository_view_list",
                         login_required(self.repository_view_list), methods=['get'])
        app.add_url_rule('/add_list', "add_list", login_required(self.add_list), methods=['get', "post"])
        app.add_url_rule('/delete_list/<list_name>', "delete_list", login_required(self.delete_list), methods=['post'])
        app.add_url_rule('/create_duplicate_list', "create_duplicate_list",
                         login_required(self.create_duplicate_list), methods=['get', 'post'])

    def view_list(self, list_name):
        the_list = current_user.get_list(list_name)
        lstring = ""
        for w in the_list:
            lstring += w + "\n"
        return render_template("user_manage/list_viewer.html",
                               list_name=list_name,
                               the_list_as_string=lstring,
                               read_only_string="")

    def repository_view_list(self, list_name):
        the_list = repository_user.get_list(list_name)
        lstring = ""
        for w in the_list:
            lstring += w + "\n"
        return render_template("user_manage/list_viewer.html",
                               list_name=list_name,
                               the_list_as_string=lstring,
                               read_only_string="readonly")

    def grab_metadata(self, res_name):
        if self.is_repository:
            user_obj = repository_user
        else:
            user_obj = current_user
        doc = db[user_obj.list_collection_name].find_one({self.name_field: res_name})
        if "metadata" in doc:
            mdata = doc["metadata"]
        else:
            mdata = None
        return mdata

    def save_metadata(self, res_name, tags, notes):
        doc = db[current_user.list_collection_name].find_one({"list_name": res_name})
        if "metadata" in doc:
            mdata = doc["metadata"]
        else:
            mdata = {}
        mdata["tags"] = tags
        mdata["notes"] = notes
        db[current_user.list_collection_name].update_one({"list_name": res_name}, {'$set': {"metadata": mdata}})
        self.update_selector_list()

    def add_list(self):
        user_obj = current_user
        the_file = request.files['file']
        if db[user_obj.list_collection_name].find_one({"list_name": the_file.filename}) is not None:
            return jsonify({"success": False, "alert_type": "alert-warning",
                            "message": "A list with that name already exists"})
        the_list = load_a_list(the_file)
        metadata = global_tile_manager.create_initial_metadata()
        data_dict = {"list_name": the_file.filename, "the_list": the_list, "metadata": metadata}
        db[user_obj.list_collection_name].insert_one(data_dict)
        self.update_selector_list(select=the_file.filename)
        return jsonify({"success": True})

    def delete_list(self, list_name):
        user_obj = current_user
        db[user_obj.list_collection_name].delete_one({"list_name": list_name})
        self.update_selector_list()
        return jsonify({"success": True})

    def create_duplicate_list(self):
        user_obj = current_user
        list_to_copy = request.json['res_to_copy']
        new_list_name = request.json['new_res_name']
        if db[user_obj.list_collection_name].find_one({"list_name": new_list_name}) is not None:
            return jsonify({"success": False, "alert_type": "alert-warning",
                            "message": "A list with that name already exists"})
        old_list_dict = db[user_obj.list_collection_name].find_one({"list_name": list_to_copy})
        metadata = global_tile_manager.create_initial_metadata()
        new_list_dict = {"list_name": new_list_name, "the_list": old_list_dict["the_list"], "metadata": metadata}
        db[user_obj.list_collection_name].insert_one(new_list_dict)
        self.update_selector_list(select=new_list_name)
        return jsonify({"success": True})


class RepositoryListManager(ListManager):
    rep_string = "repository-"
    is_repository = True

    def add_rules(self):
        pass


# noinspection PyMethodMayBeStatic,PyBroadException
class CollectionManager(ResourceManager):
    collection_list = "data_collections"
    collection_list_with_metadata = "data_collection_names_with_metadata"
    collection_name = ""
    name_field = ""

    def add_rules(self):
        app.add_url_rule('/main/<collection_name>', "main", login_required(self.main), methods=['get'])
        app.add_url_rule('/import_as_table/<collection_name>', "import_as_table",
                         login_required(self.import_as_table), methods=['get', "post"])
        app.add_url_rule('/import_as_freeform/<collection_name>', "import_as_freeform",
                         login_required(self.import_as_freeform), methods=['get', "post"])
        app.add_url_rule('/delete_collection/<collection_name>', "delete_collection",
                         login_required(self.delete_collection), methods=['post'])
        app.add_url_rule('/duplicate_collection', "duplicate_collection",
                         login_required(self.duplicate_collection), methods=['post', 'get'])
        app.add_url_rule('/download_collection/<collection_name>/<new_name>', "download_collection",
                         login_required(self.download_collection), methods=['post', 'get'])
        app.add_url_rule('/combine_collections/<base_collection_name>/<collection_to_add>', "combine_collections",
                         login_required(self.combine_collections), methods=['post', 'get'])

    def main(self, collection_name):
        user_obj = current_user
        cname = user_obj.build_data_collection_name(collection_name)
        main_id = create_container("tactic_main_image", owner=user_obj.get_id())
        caddress = get_address(main_id, "bridge")
        send_direct_request_to_container(tactic_app.megaplex_id, "register_container",
                                         {"container_id": main_id, "address": caddress})
        global_tile_manager.add_user(user_obj.username)

        the_collection = db[cname]
        mdata = the_collection.find_one({"name": "__metadata__"})
        if "type" in mdata and mdata["type"] == "freeform":
            doc_type = "freeform"
        else:
            doc_type = "table"

        data_dict = {"collection_name": cname,
                     "main_id": main_id,
                     "user_id": current_user.get_id(),
                     "megaplex_address": tactic_app.megaplex_address,
                     "project_collection_name": user_obj.project_collection_name,
                     "mongo_uri": mongo_uri,
                     "doc_type": doc_type,
                     "base_figure_url": url_for("figure_source", tile_id="tile_id", figure_name="X")[:-1]}

        result = send_direct_request_to_container(main_id, "initialize_mainwindow", data_dict).json()
        if not result["success"]:
            return result["message_string"]
        short_collection_name = re.sub("^.*?\.data_collection\.", "", collection_name)

        doc_names = []

        for f in the_collection.find():
            fname = f["name"].encode("ascii", "ignore")
            if fname == "__metadata__":
                continue
            else:
                doc_names.append(fname)

        doc_names.sort()
        return render_template("main.html",
                               collection_name=cname,
                               window_title=short_collection_name,
                               project_name='',
                               main_id=main_id,
                               doc_names=doc_names,
                               use_ssl=str(use_ssl),
                               console_html="",
                               is_table=(doc_type == "table"),
                               short_collection_name=short_collection_name,
                               new_tile_info="",
                               uses_codemirror="True")

    def download_collection(self, collection_name, new_name):
        user_obj = current_user
        cname = user_obj.build_data_collection_name(collection_name)
        the_collection = db[cname]

        wb = openpyxl.Workbook()
        first = True
        for f in the_collection.find():
            doc_name = f["name"].encode("ascii", "ignore")
            if doc_name == "__metadata__":
                continue
            sheet_name = re.sub(r"[\[\]\*\/\\ \?\:]", r"-", doc_name)[:25]
            if first:
                ws = wb.active
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(title=sheet_name)
            data_rows = f["data_rows"]
            header_list = f["header_list"]
            for c, header in enumerate(header_list, start=1):
                _ = ws.cell(row=1, column=c, value=header)
            for r, row in enumerate(data_rows.values(), start=2):
                for c, header in enumerate(header_list, start=1):
                    _ = ws.cell(row=r, column=c, value=row[header])
            # noinspection PyUnresolvedReferences
        virtual_notebook = openpyxl.writer.excel.save_virtual_workbook(wb)
        str_io = cStringIO.StringIO()
        str_io.write(virtual_notebook)
        str_io.seek(0)
        return send_file(str_io,
                         attachment_filename=new_name,
                         as_attachment=True)

    def grab_metadata(self, res_name):
        if self.is_repository:
            user_obj = repository_user
        else:
            user_obj = current_user
        cname = user_obj.build_data_collection_name(res_name)
        mdata = db[cname].find_one({"name": "__metadata__"})
        return mdata

    def save_metadata(self, res_name, tags, notes):
        cname = current_user.build_data_collection_name(res_name)
        mdata = db[cname].find_one({"name": "__metadata__"})
        if mdata is None:
            db[cname].insert_one({"name": "__metadata__", "tags": tags, "notes": notes})
        else:
            db[cname].update_one({"name": "__metadata__"},
                                 {'$set': {"tags": tags, "notes": notes}})
        self.update_selector_list()

    def autosplit_doc(self, filename, full_dict):
        sorted_int_keys = sorted([int(key) for key in full_dict.keys()])
        counter = 0
        doc_list = []
        doc_rows = {}
        doc_counter = 1
        for r in sorted_int_keys:
            doc_rows[str(counter)] = full_dict[str(r)]
            counter += 1
            if counter == AUTOSPLIT_SIZE:
                doc_list.append({
                    "name": filename + "__" + str(doc_counter),
                    "data_rows": doc_rows
                })
                counter = 0
                doc_rows = {}
                doc_counter += 1
        if counter > 0:
            doc_list.append({
                "name": filename + "__" + str(doc_counter),
                "data_rows": doc_rows
            })
        return doc_list

    def combine_collections(self, base_collection_name, collection_to_add):
        try:
            user_obj = current_user
            base_full_name = user_obj.build_data_collection_name(base_collection_name)
            add_full_name = user_obj.build_data_collection_name(collection_to_add)
            base_collection = db[base_full_name]
            add_collection = db[add_full_name]
            for f in add_collection.find():
                fname = f["name"].encode("ascii", "ignore")
                if fname == "__metadata__":
                    continue
                base_collection.insert_one(f)
            db[base_full_name].update_one({"name": "__metadata__"},
                                          {'$set': {"updated": datetime.datetime.today()}})
            self.update_selector_list(base_collection_name)
            return jsonify({"message": "Collections successfull combined", "alert_type": "alert-success"})
        except:
            error_string = "Error combining collections: " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
            return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})

    def import_as_table(self, collection_name):
        user_obj = current_user
        file_list = request.files.getlist("file")
        full_collection_name = user_obj.build_data_collection_name(collection_name)
        if full_collection_name in db.collection_names():
            return jsonify({"success": False, "message": "There is already a collection with that name.",
                            "alert_type": "alert-warning"})
        mdata = global_tile_manager.create_initial_metadata()
        try:
            db[full_collection_name].insert_one({"name": "__metadata__", "datetime": mdata["datetime"],
                                                 "updated": mdata["updated"], "tags": "", "notes": "", "type": "table"})
        except:
            error_string = "Error creating collection: " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
            return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})

        for the_file in file_list:
            filename, file_extension = os.path.splitext(the_file.filename)
            filename = filename.encode("ascii", "ignore")
            if file_extension == ".csv":
                (success, result_dict, header_list) = read_csv_file_to_dict(the_file)
            elif file_extension == ".tsv":
                (success, result_dict, header_list) = read_tsv_file_to_dict(the_file)
            elif file_extension == ".txt":
                (success, result_dict, header_list) = read_txt_file_to_dict(the_file)
            # elif file_extension == ".xml":
            #     (success, result_dict, header_list) = read_xml_file_to_dict(file)
            else:
                return jsonify({"success": False, "message": "Not a valid file extension " + file_extension,
                                "alert_type": "alert-warning"})
            if not success:  # then result_dict contains an error object
                e = result_dict
                return jsonify({"message": e.message, "alert_type": "alert-danger"})

            try:
                if AUTOSPLIT and len(result_dict.keys()) > AUTOSPLIT_SIZE:
                    docs = self.autosplit_doc(filename, result_dict)
                    for doc in docs:
                        db[full_collection_name].insert_one({"name": doc["name"], "data_rows": doc["data_rows"],
                                                             "header_list": header_list, "type": "table"})
                else:
                    db[full_collection_name].insert_one({"name": filename, "data_rows": result_dict,
                                                         "header_list": header_list})
            except:
                error_string = "Error creating collection: " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
                stop_spinner()
                return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})

        self.update_selector_list(collection_name)
        return jsonify({"message": "Collection successfully loaded", "alert_type": "alert-success"})

    def import_as_freeform(self, collection_name):
        user_obj = current_user
        file_list = request.files.getlist("file")
        full_collection_name = user_obj.build_data_collection_name(collection_name)
        if full_collection_name in db.collection_names():
            return jsonify({"success": False, "message": "There is already a collection with that name.",
                            "alert_type": "alert-warning"})
        mdata = global_tile_manager.create_initial_metadata()
        try:
            db[full_collection_name].insert_one({"name": "__metadata__", "datetime": mdata["datetime"],
                                                 "updated": mdata["updated"], "tags": "",
                                                 "notes": "", "type": "freeform"})
        except:
            error_string = "Error creating collection: " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
            return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})

        for the_file in file_list:
            filename, file_extension = os.path.splitext(the_file.filename)
            filename = filename.encode("ascii", "ignore")
            (success, result_txt) = read_freeform_file(the_file)
            if not success:  # then result_dict contains an error object
                e = result_txt
                return jsonify({"message": e.message, "alert_type": "alert-danger"})

            try:
                save_dict = {"name": filename, "file_id": fs.put(result_txt)}
                db[full_collection_name].insert_one(save_dict)
            except:
                error_string = "Error creating collection: " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
                stop_spinner()
                return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})

        self.update_selector_list(collection_name)
        return jsonify({"message": "Collection successfully loaded", "alert_type": "alert-success"})

    def delete_collection(self, collection_name):
        user_obj = current_user
        result = user_obj.remove_collection(collection_name)
        self.update_selector_list()
        return jsonify({"success": result})

    def duplicate_collection(self):
        user_obj = current_user
        collection_to_copy = user_obj.full_collection_name(request.json['res_to_copy'])
        new_collection_name = user_obj.full_collection_name(request.json['new_res_name'])
        if new_collection_name in db.collection_names():
            return jsonify({"success": False, "message": "There is already a collection with that name.",
                            "alert_type": "alert-warning"})

        for doc in db[collection_to_copy].find():
            if "file_id" in doc:
                doc_text = fs.get(doc["file_id"]).read()
                doc["file_id"] = fs.put(doc_text)
            db[new_collection_name].insert_one(doc)
        self.update_selector_list(request.json['new_res_name'])
        return jsonify({"success": True})


class RepositoryCollectionManager(CollectionManager):
    rep_string = "repository-"
    is_repository = True

    def add_rules(self):
        pass


class ProjectManager(ResourceManager):
    collection_list = "project_names"
    collection_list_with_metadata = "project_names_with_metadata"
    collection_name = "project_collection_name"
    name_field = "project_name"

    def add_rules(self):
        app.add_url_rule('/delete_project/<project_name>', "delete_project", login_required(self.delete_project),
                         methods=['post'])

    def delete_project(self, project_name):
        current_user.remove_project(project_name)
        self.update_selector_list()
        return jsonify({"success": True})

    def grab_metadata(self, res_name):
        if self.is_repository:
            user_obj = repository_user
        else:
            user_obj = current_user
        doc = db[user_obj.project_collection_name].find_one({self.name_field: res_name})
        if "metadata" in doc:
            mdata = doc["metadata"]
        else:
            mdata = None
        return mdata

    def save_metadata(self, res_name, tags, notes):
        doc = db[current_user.project_collection_name].find_one({"project_name": res_name})
        if "metadata" in doc:
            mdata = doc["metadata"]
        else:
            mdata = {}
        mdata["tags"] = tags
        mdata["notes"] = notes
        db[current_user.project_collection_name].update_one({"project_name": res_name}, {'$set': {"metadata": mdata}})
        self.update_selector_list()


class RepositoryProjectManager(ProjectManager):
    rep_string = "repository-"
    is_repository = True

    def add_rules(self):
        pass


# noinspection PyMethodMayBeStatic,PyBroadException
class TileManager(ResourceManager):
    collection_list = "tile_module_names"
    collection_list_with_metadata = "tile_module_names_with_metadata"
    collection_name = "tile_collection_name"
    name_field = "tile_module_name"

    def add_rules(self):
        app.add_url_rule('/view_module/<module_name>', "view_module",
                         login_required(self.view_module), methods=['get'])
        app.add_url_rule('/get_module_code/<module_name>', "get_module_code",
                         login_required(self.get_module_code), methods=['get', 'post'])
        app.add_url_rule('/view_in_creator/<module_name>', "view_in_creator",
                         login_required(self.view_in_creator), methods=['get'])
        app.add_url_rule('/last_saved_view/<module_name>', "last_saved_view",
                         login_required(self.last_saved_view), methods=['get'])
        app.add_url_rule('/repository_view_module/<module_name>', "repository_view_module",
                         login_required(self.repository_view_module), methods=['get'])
        app.add_url_rule('/load_tile_module/<tile_module_name>', "load_tile_module",
                         login_required(self.load_tile_module), methods=['get', 'post'])
        app.add_url_rule('/unload_all_tiles', "unload_all_tiles",
                         login_required(self.unload_all_tiles), methods=['get', 'post'])
        app.add_url_rule('/add_tile_module', "add_tile_module",
                         login_required(self.add_tile_module), methods=['get', "post"])
        app.add_url_rule('/delete_tile_module/<tile_module_name>', "delete_tile_module",
                         login_required(self.delete_tile_module), methods=['post'])
        app.add_url_rule('/create_tile_module', "create_tile_module",
                         login_required(self.create_tile_module), methods=['get', 'post'])
        app.add_url_rule('/request_update_loaded_tile_list', "request_update_loaded_tile_list",
                         login_required(self.request_update_loaded_tile_list), methods=['get', 'post'])
        app.add_url_rule('/create_duplicate_tile', "create_duplicate_tile",
                         login_required(self.create_duplicate_tile), methods=['get', 'post'])

    def grab_metadata(self, res_name):
        if self.is_repository:
            user_obj = repository_user
        else:
            user_obj = current_user
        doc = db[user_obj.tile_collection_name].find_one({self.name_field: res_name})
        if "metadata" in doc:
            mdata = doc["metadata"]
        else:
            mdata = None
        return mdata

    def save_metadata(self, res_name, tags, notes):
        doc = db[current_user.tile_collection_name].find_one({"tile_module_name": res_name})
        if "metadata" in doc:
            mdata = doc["metadata"]
        else:
            mdata = {}
        mdata["tags"] = tags
        mdata["notes"] = notes
        db[current_user.tile_collection_name].update_one({"tile_module_name": res_name}, {'$set': {"metadata": mdata}})
        self.update_selector_list()

    def view_module(self, module_name):
        return render_template("user_manage/module_viewer.html",
                               module_name=module_name,
                               read_only_string="",
                               api_html=api_html,
                               uses_codemirror="True")

    def get_module_code(self, module_name):
        user_obj = current_user
        module_code = user_obj.get_tile_module(module_name)
        return jsonify({"success": True, "module_code": module_code})

    def last_saved_view(self, module_name):
        tile_dict = current_user.get_tile_dict(module_name)
        if "last_saved" in tile_dict and tile_dict["last_saved"] == "creator":
            result = self.view_in_creator(module_name)
        else:
            result = self.view_module(module_name)
        return result


    def view_in_creator(self, module_name):
        option_types = [{"name": "text"},
                        {"name": "int"},
                        {"name": "boolean"},
                        {"name": "textarea"},
                        {"name": "codearea"},
                        {"name": "document_select"},
                        {"name": "list_select"},
                        {"name": "palette_select"},
                        {"name": "pipe_select"},
                        {"name": "custom_list"},
                        {"name": "function_select"},
                        {"name": "class_select"},
                        {"name": "pipe_select"}]
        revised_api_dlist = []
        for cat in ordered_api_categories:
            the_list = api_dict_by_category[cat]
            new_list = []
            for api_item in the_list:
                new_list.append({"name": api_item["name"]})
            if len(new_list) > 0:
                revised_api_dlist.append({"cat_name": cat, "cat_list": new_list})
        return render_template("user_manage/tile_creator.html",
                               module_name=module_name,
                               read_only_string="",
                               api_html=api_html,
                               use_ssl=use_ssl,
                               api_dlist=revised_api_dlist,
                               option_types=option_types,
                               uses_codemirror="True")

    def repository_view_module(self, module_name):
        user_obj = repository_user
        module_code = user_obj.get_tile_module(module_name)
        return render_template("user_manage/module_viewer.html",
                               module_name=module_name,
                               module_code=module_code,
                               read_only_string="readonly",
                               api_html=api_html,
                               uses_codemirror=True)

    def load_tile_module(self, tile_module_name, return_json=True, user_obj=None):
        try:
            if user_obj is None:
                user_obj = current_user
            tile_module = user_obj.get_tile_module(tile_module_name)

            result = send_direct_request_to_container(global_tile_manager.test_tile_container_id, "load_source",
                                                      {"tile_code": tile_module,
                                                       "megaplex_address": tactic_app.megaplex_address})
            res_dict = result.json()

            if not res_dict["success"]:
                return jsonify({"success": False, "message": res_dict["message_string"], "alert_type": "alert-warning"})
            category = res_dict["category"]

            global_tile_manager.add_user_tile_module(user_obj.username,
                                                     category,
                                                     res_dict["tile_name"],
                                                     tile_module,
                                                     tile_module_name)

            socketio.emit('update-loaded-tile-list', {"html": self.render_loaded_tile_list(user_obj)},
                          namespace='/user_manage', room=user_obj.get_id())
            socketio.emit('update-menus', {}, namespace='/main', room=user_obj.get_id())
            if return_json:
                return jsonify({"success": True, "message": "Tile module successfully loaded",
                                "alert_type": "alert-success"})
            else:
                return {"success": True}
        except:
            error_string = "Error loading tile: " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
            if return_json:
                return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})
            else:
                return {"success": False, "message": error_string}

    def unload_all_tiles(self):
        try:
            global_tile_manager.unload_user_tiles(current_user.username)
            socketio.emit('update-loaded-tile-list', {"html": self.render_loaded_tile_list()},
                          namespace='/user_manage', room=current_user.get_id())
            socketio.emit('update-menus', {}, namespace='/main', room=current_user.get_id())
            return jsonify({"message": "Tiles successfully unloaded", "alert_type": "alert-success"})
        except:
            error_string = "Error unloading tiles: " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
            return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})

    def add_tile_module(self):
        user_obj = current_user
        f = request.files['file']
        if db[user_obj.tile_collection_name].find_one({"tile_module_name": f.filename}) is not None:
            return jsonify({"success": False, "alert_type": "alert-warning",
                            "message": "A module with that name already exists"})
        the_module = f.read()
        metadata = global_tile_manager.create_initial_metadata()
        data_dict = {"tile_module_name": f.filename, "tile_module": the_module, "metadata": metadata}
        db[user_obj.tile_collection_name].insert_one(data_dict)
        self.update_selector_list(f.filename)
        return jsonify({"success": True})

    def create_duplicate_tile(self):
        user_obj = current_user
        tile_to_copy = request.json['res_to_copy']
        new_tile_name = request.json['new_res_name']
        if db[user_obj.tile_collection_name].find_one({"tile_module_name": new_tile_name}) is not None:
            return jsonify({"success": False, "alert_type": "alert-warning",
                            "message": "A tile with that name already exists"})
        old_tile_dict = db[user_obj.tile_collection_name].find_one({"tile_module_name": tile_to_copy})
        metadata = global_tile_manager.create_initial_metadata()
        new_tile_dict = {"tile_module_name": new_tile_name, "tile_module": old_tile_dict["tile_module"],
                         "metadata": metadata}
        db[user_obj.tile_collection_name].insert_one(new_tile_dict)
        self.update_selector_list(select=new_tile_name)
        return jsonify({"success": True})

    def create_tile_module(self):
        user_obj = current_user
        new_tile_name = request.json['new_res_name']
        template_name = request.json["template_name"]
        last_saved = request.json["last_saved"]
        if db[user_obj.tile_collection_name].find_one({"tile_module_name": new_tile_name}) is not None:
            return jsonify({"success": False, "alert_type": "alert-warning",
                            "message": "A module with that name already exists"})
        mongo_dict = db["repository.tiles"].find_one({"tile_module_name": template_name})
        template = mongo_dict["tile_module"]

        metadata = global_tile_manager.create_initial_metadata()
        data_dict = {"tile_module_name": new_tile_name, "tile_module": template, "metadata": metadata,
                     "last_saved": last_saved}
        db[current_user.tile_collection_name].insert_one(data_dict)
        self.update_selector_list(new_tile_name)
        return jsonify({"success": True})

    def delete_tile_module(self, tile_module_name):
        user_obj = current_user
        db[user_obj.tile_collection_name].delete_one({"tile_module_name": tile_module_name})
        self.update_selector_list()
        return jsonify({"success": True})

    def render_loaded_tile_list(self, user_obj=None):
        if user_obj is None:
            user_obj = current_user
        loaded_tiles = global_tile_manager.get_loaded_user_tiles_list(user_obj.username)
        with app.test_request_context():
            result = render_template("user_manage/loaded_tile_list.html", user_tile_name_list=loaded_tiles)
        return result

    def request_update_loaded_tile_list(self):
        return self.render_loaded_tile_list()


class RepositoryTileManager(TileManager):
    rep_string = "repository-"
    is_repository = True

    def add_rules(self):
        pass


class CodeManager(ResourceManager):
    collection_list = "code_names"
    collection_list_with_metadata = "code_names_with_metadata"
    collection_name = "code_collection_name"
    name_field = "code_name"

    def add_rules(self):
        app.add_url_rule('/view_code/<code_name>', "view_code",
                         login_required(self.view_code), methods=['get'])
        app.add_url_rule('/repository_view_code/<code_name>', "repository_view_code",
                         login_required(self.repository_view_code), methods=['get'])
        app.add_url_rule('/add_code', "add_code",
                         login_required(self.add_code), methods=['get', "post"])
        app.add_url_rule('/delete_code/<code_name>', "delete_code",
                         login_required(self.delete_code), methods=['post'])
        app.add_url_rule('/create_code', "create_code",
                         login_required(self.create_code), methods=['get', 'post'])
        app.add_url_rule('/create_duplicate_code', "create_duplicate_code",
                         login_required(self.create_duplicate_code), methods=['get', 'post'])

    def grab_metadata(self, res_name):
        if self.is_repository:
            user_obj = repository_user
        else:
            user_obj = current_user
        doc = db[user_obj.code_collection_name].find_one({self.name_field: res_name})
        if "metadata" in doc:
            mdata = doc["metadata"]
        else:
            mdata = None
        return mdata

    def save_metadata(self, res_name, tags, notes):
        doc = db[current_user.code_collection_name].find_one({"code_name": res_name})
        if "metadata" in doc:
            mdata = doc["metadata"]
        else:
            mdata = {}
        mdata["tags"] = tags
        mdata["notes"] = notes
        db[current_user.code_collection_name].update_one({"code_name": res_name}, {'$set': {"metadata": mdata}})
        self.update_selector_list()

    def view_code(self, code_name):
        user_obj = current_user
        the_code = user_obj.get_code(code_name)
        return render_template("user_manage/code_viewer.html",
                               code_name=code_name,
                               the_code=the_code,
                               read_only_string="",
                               uses_codemirror="True")

    def repository_view_code(self, code_name):
        user_obj = repository_user
        the_code = user_obj.get_code(code_name)
        return render_template("user_manage/code_viewer.html",
                               code_name=code_name,
                               the_code=the_code,
                               read_only_string="readonly",
                               uses_codemirror="True")

    def load_code(self, the_code):
        result = send_direct_request_to_container(global_tile_manager.test_tile_container_id, "clear_and_load_code",
                                                  {"the_code": the_code,
                                                   "megaplex_address": tactic_app.megaplex_address})
        res_dict = result.json()
        return res_dict

    def add_code(self):
        user_obj = current_user
        f = request.files['file']
        if db[user_obj.code_collection_name].find_one({"code_name": f.filename}) is not None:
            return jsonify({"success": False, "alert_type": "alert-warning",
                            "message": "A code resource with that name already exists"})
        the_code = f.read()
        metadata = global_tile_manager.create_initial_metadata()

        load_result = self.load_code(the_code)
        if not load_result["success"]:
            return jsonify(load_result)

        metadata["classes"] = load_result["classes"]
        metadata["functions"] = load_result["functions"]

        data_dict = {"code_name": f.filename, "the_code": the_code, "metadata": metadata}
        db[user_obj.code_collection_name].insert_one(data_dict)
        self.update_selector_list(f.filename)

        return jsonify({"success": True})

    def create_duplicate_code(self):
        user_obj = current_user
        code_to_copy = request.json['res_to_copy']
        new_code_name = request.json['new_res_name']
        if db[user_obj.code_collection_name].find_one({"code_name": new_code_name}) is not None:
            return jsonify({"success": False, "alert_type": "alert-warning",
                            "message": "A code resource with that name already exists"})
        old_code_dict = db[user_obj.code_collection_name].find_one({"code_name": code_to_copy})
        metadata = global_tile_manager.create_initial_metadata()
        metadata["classes"] = old_code_dict["metadata"]["classes"]
        metadata["functions"] = old_code_dict["metadata"]["functions"]
        new_code_dict = {"code_name": new_code_name, "the_code": old_code_dict["the_code"], "metadata": metadata}
        db[user_obj.code_collection_name].insert_one(new_code_dict)
        self.update_selector_list(select=new_code_name)
        return jsonify({"success": True})

    def create_code(self):
        user_obj = current_user
        new_code_name = request.json['new_res_name']
        template_name = request.json["template_name"]
        if db[user_obj.code_collection_name].find_one({"code_name": new_code_name}) is not None:
            return jsonify({"success": False, "alert_type": "alert-warning",
                            "message": "A module with that name already exists"})
        mongo_dict = db[repository_user.code_collection_name].find_one({"code_name": template_name})
        template = mongo_dict["the_code"]

        metadata = global_tile_manager.create_initial_metadata()
        metadata["functions"] = []
        metadata["classes"] = []
        data_dict = {"code_name": new_code_name, "the_code": template, "metadata": metadata}
        db[current_user.code_collection_name].insert_one(data_dict)
        self.update_selector_list(new_code_name)
        return jsonify({"success": True})

    def delete_code(self, code_name):
        user_obj = current_user
        db[user_obj.code_collection_name].delete_one({"code_name": code_name})
        self.update_selector_list()
        return jsonify({"success": True})


class RepositoryCodeManager(CodeManager):
    rep_string = "repository-"
    is_repository = True

    def add_rules(self):
        pass

repository_user = User.get_user_by_username("repository")

list_manager = ListManager("list")
repository_list_manager = RepositoryListManager("list")

collection_manager = CollectionManager("collection")
repository_collection_manager = RepositoryCollectionManager("collection")

project_manager = ProjectManager("project")
repository_project_manager = RepositoryProjectManager("project")

tile_manager = TileManager("tile")
repository_tile_manager = RepositoryTileManager("tile")

code_manager = CodeManager("code")
repository_code_manager = RepositoryCodeManager("code")

managers = {
    "list": [list_manager, repository_list_manager],
    "collection": [collection_manager, repository_collection_manager],
    "project": [project_manager, repository_project_manager],
    "tile": [tile_manager, repository_tile_manager],
    "code": [code_manager, repository_code_manager]
}


@app.route('/user_manage')
@login_required
def user_manage():
    return render_template('user_manage/user_manage.html',
                           use_ssl=str(use_ssl))

# Metadata views


@app.route('/grab_metadata', methods=['POST'])
@login_required
def grab_metadata():
    try:
        res_type = request.json["res_type"]
        res_name = request.json["res_name"]
        manager = get_manager_for_type(res_type)
        mdata = manager.grab_metadata(res_name)
        if mdata is None:
            return jsonify({"success": False, "message": "No metadata found", "alert_type": "alert-warning"})
        else:
            if "datetime" in mdata:
                datestring = mdata["datetime"].strftime("%b %d, %Y, %H:%M")
            else:
                datestring = ""
            return jsonify({"success": True, "datestring": datestring, "tags": mdata["tags"], "notes": mdata["notes"]})
    except:
        error_string = "Error getting metadata: " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})


@app.route('/grab_repository_metadata', methods=['POST'])
@login_required
def grab_repository_metadata():
    try:
        res_type = request.json["res_type"]
        res_name = request.json["res_name"]
        manager = get_manager_for_type(res_type, is_repository=True)
        mdata = manager.grab_metadata(res_name)
        if mdata is None:
            return jsonify({"success": False, "message": "No repository metadata found", "alert_type": "alert-warning"})
        else:
            if "datetime" in mdata:
                datestring = mdata["datetime"].strftime("%b %d, %Y, %H:%M")
            else:
                datestring = ""
            return jsonify({"success": True, "datestring": datestring, "tags": mdata["tags"], "notes": mdata["notes"]})
    except:
        error_string = "Error getting metadata: " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})


@app.route('/save_metadata', methods=['POST'])
@login_required
def save_metadata():
    try:
        res_type = request.json["res_type"]
        res_name = request.json["res_name"]
        tags = request.json["tags"]
        notes = request.json["notes"]
        manager = get_manager_for_type(res_type)
        manager.save_metadata(res_name, tags, notes)
        return jsonify({"success": True, "message": "Saved metadata", "alert_type": "alert-success"})
    except:
        error_string = "Error saving metadata: " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})


@app.route('/search_resource', methods=['POST'])
@login_required
def search_resource():
    txt = request.json["text"]
    res_type = request.json["res_type"]
    search_type = request.json["search_type"]
    search_location = request.json["location"]
    if search_location == "repository":
        user_obj = repository_user
        manager = get_manager_for_type(res_type, is_repository=True)
    else:
        user_obj = current_user
        manager = get_manager_for_type(res_type)
    if search_type == "search":
        the_list = user_obj.get_resource_names(res_type, search_filter=txt)
    else:
        the_list = user_obj.get_resource_names(res_type, tag_filter=txt)

    res_array = manager.build_resource_array(the_list)
    result = manager.build_html_table_from_data_list(res_array)
    return jsonify({"html": result})


indent_unit = "    "
def remove_indents(the_str, number_indents):
    total_indent = indent_unit * number_indents
    result = re.sub(r"\n" + total_indent, "\n", the_str)
    result = re.sub(r"^" + total_indent, "", result)
    return result

def insert_indents(the_str, number_indents):
    total_indent = indent_unit * number_indents
    result = re.sub(r"\n", r"\n" + total_indent, the_str)
    result = total_indent + result
    return result

def build_code(data_dict):
    export_list = data_dict["exports"]
    export_list_of_dicts = [{"name": exp_name} for exp_name in export_list]
    extra_methods = insert_indents(data_dict["extra_methods"], 1)
    render_content_body = insert_indents(data_dict["render_content_body"], 2)
    if data_dict["is_mpl"]:
        draw_plot_body = insert_indents(data_dict["draw_plot_body"], 2)
    else:
        draw_plot_body = ""
    options = data_dict["options"]
    for opt_dict in options:
        if "default" not in opt_dict:
            opt_dict["default"] = "None"
        elif isinstance(opt_dict["default"], basestring):
            opt_dict["default"] = '"' + opt_dict["default"] + '"'
        opt_dict["default"] = str(opt_dict["default"])
        if "special_list" in opt_dict:
            opt_dict["special_list"] = "[" + opt_dict["special_list"] + "]"
    full_code = render_template("user_manage/tile_creator_template",
                                class_name=data_dict["module_name"],
                                category=data_dict["category"],
                                exports=export_list_of_dicts,
                                options=data_dict["options"],
                                is_mpl=data_dict["is_mpl"],
                                extra_methods=extra_methods,
                                render_content_body=render_content_body,
                                draw_plot_body=draw_plot_body)
    return full_code

@app.route('/update_module', methods=['post'])
@login_required
def update_module():
    try:
        data_dict = request.json
        module_name = data_dict["module_name"]
        last_saved = data_dict["last_saved"]
        if last_saved == "viewer":
            module_code = data_dict["new_code"]
        else:
            module_code = build_code(data_dict)
        doc = db[current_user.tile_collection_name].find_one({"tile_module_name": module_name})
        if "metadata" in doc:
            mdata = doc["metadata"]
        else:
            mdata = {}
        mdata["tags"] = data_dict["tags"]
        mdata["notes"] = data_dict["notes"]
        mdata["updated"] = datetime.datetime.today()


        db[current_user.tile_collection_name].update_one({"tile_module_name": module_name},
                                                         {'$set': {"tile_module": module_code, "metadata": mdata,
                                                                   "last_saved": last_saved}})
        tile_manager.update_selector_list()
        return jsonify({"success": True, "message": "Module Successfully Saved", "alert_type": "alert-success"})
    except:
        error_string = "Error saving module " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})


@app.route('/rename_module/<old_name>', methods=['post'])
@login_required
def rename_module(old_name):
    try:
        new_name = request.json["new_name"]
        db[current_user.tile_collection_name].update_one({"tile_module_name": old_name},
                                                         {'$set': {"tile_module_name": new_name}})
        tile_manager.update_selector_list()
        return jsonify({"success": True, "message": "Module Successfully Saved", "alert_type": "alert-success"})
    except:
        error_string = "Error renaming module " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})


@app.route('/rename_list/<old_name>', methods=['post'])
@login_required
def rename_list(old_name):
    try:
        new_name = request.json["new_name"]
        db[current_user.list_collection_name].update_one({"list_name": old_name},
                                                         {'$set': {"list_name": new_name}})
        list_manager.update_selector_list()
        return jsonify({"success": True, "message": "List name changed", "alert_type": "alert-success"})
    except:
        error_string = "Error renaming list " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})


@app.route('/rename_code/<old_name>', methods=['post'])
@login_required
def rename_code(old_name):
    try:
        new_name = request.json["new_name"]
        db[current_user.code_collection_name].update_one({"code_name": old_name},
                                                         {'$set': {"code_name": new_name}})
        code_manager.update_selector_list()
        return jsonify({"success": True, "message": "Module Successfully Saved", "alert_type": "alert-success"})
    except:
        error_string = "Error renaming module " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})


@app.route('/update_code', methods=['post'])
@login_required
def update_code():
    try:
        data_dict = request.json
        code_name = data_dict["code_name"]
        the_code = data_dict["new_code"]
        doc = db[current_user.code_collection_name].find_one({"code_name": code_name})
        if "metadata" in doc:
            mdata = doc["metadata"]
        else:
            mdata = {}
        mdata["tags"] = data_dict["tags"]
        mdata["notes"] = data_dict["notes"]
        mdata["updated"] = datetime.datetime.today()

        load_result = code_manager.load_code(the_code)
        if not load_result["success"]:
            return jsonify(load_result)

        mdata["classes"] = load_result["classes"]
        mdata["functions"] = load_result["functions"]

        db[current_user.code_collection_name].update_one({"code_name": code_name},
                                                         {'$set': {"the_code": the_code, "metadata": mdata}})
        code_manager.update_selector_list()
        return jsonify({"success": True, "message": "Module Successfully Saved", "alert_type": "alert-success"})
    except:
        error_string = "Error saving code resource " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})


@app.route('/update_list', methods=['post'])
@login_required
def update_list():
    try:
        data_dict = request.json
        list_name = data_dict["list_name"]
        new_list_as_string = data_dict["new_list_as_string"]
        new_list = new_list_as_string.split("\n")
        doc = db[current_user.list_collection_name].find_one({"list_name": list_name})
        if "metadata" in doc:
            mdata = doc["metadata"]
        else:
            mdata = {}
        mdata["tags"] = data_dict["tags"]
        mdata["notes"] = data_dict["notes"]
        mdata["updated"] = datetime.datetime.today()

        db[current_user.list_collection_name].update_one({"list_name": list_name},
                                                         {'$set': {"the_list": new_list, "metadata": mdata}})
        list_manager.update_selector_list()
        return jsonify({"success": True, "message": "List Successfully Saved", "alert_type": "alert-success"})
    except:
        error_string = "Error saving list " + str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})


@app.route('/get_modal_template', methods=['get'])
def get_modal_template():
    return send_file("templates/modal_text_request_template.html")


@app.route('/get_resource_module_template', methods=['get'])
@login_required
def get_resource_module_template():
    return send_file("templates/resource_module_template.html")


@socketio.on('connect', namespace='/user_manage')
@login_required
def connected_msg():
    print"client connected"


@socketio.on('join', namespace='/user_manage')
@login_required
def on_join(data):
    room = data["user_id"]
    join_room(room)
    print "user joined room " + room
    room = data["user_manage_id"]
    join_room(room)
    print "user joined room " + room
