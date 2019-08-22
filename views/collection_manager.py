
import re, datetime, sys, os
from collections import OrderedDict
from flask_login import login_required, current_user
from flask import jsonify, render_template, url_for, request, send_file
from tactic_app.users import User
from tactic_app.docker_functions import create_container, main_container_info
from tactic_app import app, db, fs, use_ssl
from tactic_app.communication_utils import make_python_object_jsonizable, debinarize_python_object
from tactic_app.communication_utils import read_temp_data, delete_temp_data
from tactic_app.mongo_accesser import MongoAccessException
import openpyxl
from openpyxl.styles import Alignment, Font
import cStringIO
import tactic_app
from tactic_app.file_handling import read_csv_file_to_list, read_tsv_file_to_list, read_txt_file_to_list
from tactic_app.file_handling import read_freeform_file, read_excel_file
from tactic_app.users import load_user

from tactic_app.resource_manager import ResourceManager, LibraryResourceManager
global_tile_manager = tactic_app.global_tile_manager
repository_user = User.get_user_by_username("repository")

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

AUTOSPLIT = False
AUTOSPLIT_SIZE = 10000

import datetime
tstring = datetime.datetime.utcnow().strftime("%Y-%H-%M-%S")


# noinspection PyMethodMayBeStatic,PyBroadException,RegExpRedundantEscape
class CollectionManager(LibraryResourceManager):
    collection_list = "data_collections"
    collection_list_with_metadata = "data_collection_names_with_metadata"
    collection_name = ""
    name_field = ""

    def add_rules(self):
        app.add_url_rule('/new_notebook', "new_notebook", login_required(self.new_notebook), methods=['get'])
        app.add_url_rule('/open_notebook/<unique_id>', "open_notebook",
                         login_required(self.open_notebook), methods=['get'])
        app.add_url_rule('/main/<collection_name>', "main", login_required(self.main), methods=['get'])
        app.add_url_rule('/import_as_table/<collection_name>/<library_id>', "import_as_table",
                         login_required(self.import_as_table), methods=['get', "post"])
        app.add_url_rule('/import_as_freeform/<collection_name>/<library_id>', "import_as_freeform",
                         login_required(self.import_as_freeform), methods=['get', "post"])
        app.add_url_rule('/delete_collection', "delete_collection",
                         login_required(self.delete_collection), methods=['post'])
        app.add_url_rule('/duplicate_collection', "duplicate_collection",
                         login_required(self.duplicate_collection), methods=['post', 'get'])
        app.add_url_rule('/download_collection/<collection_name>/<new_name>', "download_collection",
                         login_required(self.download_collection), methods=['post', 'get'])
        app.add_url_rule('/combine_collections/<base_collection_name>/<collection_to_add>', "combine_collections",
                         login_required(self.combine_collections), methods=['post', 'get'])
        app.add_url_rule('/combine_to_new_collection', "combine_to_new_collection",
                         login_required(self.combine_to_new_collection), methods=['post'])

    def new_notebook(self):
        user_obj = current_user
        main_id = main_container_info.create_main_container("new_notebook", user_obj.get_id(), user_obj.username)
        return render_template("main_notebook.html",
                               window_title="new notebook",
                               project_name='',
                               base_figure_url=url_for("figure_source", tile_id="tile_id", figure_name="X")[:-1],
                               main_id=main_id,
                               temp_data_id="",
                               use_ssl=str(use_ssl),
                               console_html="",
                               uses_codemirror="True",
                               version_string=tstring)

    def open_notebook(self, unique_id):
        the_data = read_temp_data(db, unique_id)
        user_obj = load_user(the_data["user_id"])
        main_id = main_container_info.create_main_container("new_notebook", the_data["user_id"], user_obj.username)
        return render_template("main_notebook.html",
                               window_title="new notebook",
                               project_name='',
                               base_figure_url=url_for("figure_source", tile_id="tile_id", figure_name="X")[:-1],
                               main_id=main_id,
                               temp_data_id=unique_id,
                               use_ssl=str(use_ssl),
                               console_html="",
                               uses_codemirror="True",
                               version_string=tstring)

    def main(self, collection_name):
        user_obj = current_user
        cname = user_obj.build_data_collection_name(collection_name)
        main_id = main_container_info.create_main_container(collection_name, user_obj.get_id(), user_obj.username)
        global_tile_manager.add_user(user_obj.username)

        short_collection_name = user_obj.short_collection_name(collection_name)
        mdata = user_obj.get_collection_metadata(short_collection_name)
        if "type" in mdata and mdata["type"] == "freeform":
            doc_type = "freeform"
        else:
            doc_type = "table"

        doc_names = user_obj.get_collection_docnames(short_collection_name)

        return render_template("main_react.html",
                               collection_name=cname,
                               window_title=short_collection_name,
                               project_name='',
                               base_figure_url=url_for("figure_source", tile_id="tile_id", figure_name="X")[:-1],
                               main_id=main_id,
                               main_port=main_container_info.port(main_id),
                               temp_data_id="",
                               doc_names=doc_names,
                               use_ssl=str(use_ssl),
                               console_html="",
                               is_table=(doc_type == "table"),
                               is_notebook=False,
                               is_freeform=(doc_type == 'freeform'),
                               short_collection_name=short_collection_name,
                               uses_codemirror="True",
                               version_string=tstring,
                               module_source="tactic_js/main_app.js")

    def rename_me(self, old_name):
        try:
            new_name = request.json["new_name"]
            current_user.rename_collection(old_name, new_name)
            return jsonify({"success": True, "message": "collection name changed", "alert_type": "alert-success"})
        except Exception as ex:
            return self.get_exception_for_ajax(ex, "Error renaming collection")

    def adjust_ws_col_widths(self, ws, max_col_width):
        def as_text(value):
            if value is None:
                return ""
            return unicode(value)
        for column_cells in ws.columns:
            wrap = False
            length = max(len(as_text(cell.value)) for cell in column_cells) + 5
            if length > max_col_width:
                length = max_col_width
                wrap = True
            col = ws.column_dimensions[column_cells[0].column]
            col.width = length
            if wrap:
                for cell in column_cells:
                    cell.alignment = Alignment(wrap_text=True)
                # col.alignment = Alignment(wrap_text=True)
        return

    # noinspection PyTypeChecker
    def download_collection(self, collection_name, new_name, max_col_width=50):
        user_obj = current_user
        coll_dict, doc_mdata_dict, header_list_dict, coll_mdata = user_obj.get_all_collection_info(collection_name,
                                                                                                   return_lists=False)
        wb = openpyxl.Workbook()
        first = True
        for doc_name in coll_dict.keys():
            sheet_name = re.sub(r"[\[\]\*\/\\ \?\:]", r"-", doc_name)[:25]
            if first:
                ws = wb.active
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(title=sheet_name)
            data_rows = coll_dict[doc_name]
            header_list = header_list_dict[doc_name]
            for c, header in enumerate(header_list, start=1):
                _ = ws.cell(row=1, column=c, value=header)
                ws.cell(1, c).font = Font(bold=True)
            sorted_int_keys = sorted([int(key) for key in data_rows.keys()])
            for r, _id in enumerate(sorted_int_keys, start=2):
                row = data_rows[str(_id)]
                for c, header in enumerate(header_list, start=1):
                    val = re.sub(ILLEGAL_CHARACTERS_RE, " ", unicode(row[header]))
                    _ = ws.cell(row=r, column=c, value=val)
            self.adjust_ws_col_widths(ws, max_col_width)
            # noinspection PyUnresolvedReferences
        virtual_notebook = openpyxl.writer.excel.save_virtual_workbook(wb)
        str_io = cStringIO.StringIO()
        str_io.write(virtual_notebook)
        str_io.seek(0)
        return send_file(str_io,
                         attachment_filename=new_name,
                         as_attachment=True)

    def grab_metadata(self, res_name):
        if self.is_repository:
            user_obj = repository_user
        else:
            user_obj = current_user
        return user_obj.get_collection_metadata(res_name)

    def save_metadata(self, res_name, tags, notes):
        current_user.set_collection_metadata(res_name, tags, notes)
        return

    def delete_tag(self, tag):
        cnames_with_metadata = current_user.data_collection_names_with_metadata
        for [res_name, mdata] in cnames_with_metadata:
            if mdata is None:
                continue
            tagstring = mdata["tags"]
            taglist = tagstring.split()
            if tag in taglist:
                taglist.remove(tag)
                mdata["tags"] = " ".join(taglist)
                current_user.set_collection_metadata(res_name, mdata["tags"], mdata["notes"])
        return

    def rename_tag(self, tag_changes):
        cnames_with_metadata = current_user.data_collection_names_with_metadata
        for [res_name, mdata] in cnames_with_metadata:
            tagstring = mdata["tags"]
            taglist = tagstring.split()
            for old_tag, new_tag in tag_changes:
                if old_tag in taglist:
                    taglist.remove(old_tag)
                    if new_tag not in taglist:
                        taglist.append(new_tag)
                    mdata["tags"] = " ".join(taglist)
                    current_user.set_collection_metadata(res_name, mdata["tags"], mdata["notes"])
        return

    def autosplit_doc(self, filename, full_dict):
        sorted_int_keys = sorted([int(key) for key in full_dict.keys()])
        counter = 0
        doc_list = []
        doc_rows = {}
        doc_counter = 1
        for r in sorted_int_keys:
            doc_rows[str(counter)] = full_dict[str(r)]
            counter += 1
            if counter == AUTOSPLIT_SIZE:
                doc_list.append({
                    "name": filename + "__" + str(doc_counter),
                    "data_rows": doc_rows
                })
                counter = 0
                doc_rows = {}
                doc_counter += 1
        if counter > 0:
            doc_list.append({
                "name": filename + "__" + str(doc_counter),
                "data_rows": doc_rows
            })
        return doc_list

    def combine_collections(self, base_collection_name, collection_to_add):
        try:
            user_obj = current_user

            if base_collection_name not in user_obj.data_collections:
                error_string = base_collection_name + " doesn't exist"
                return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})

            if collection_to_add not in user_obj.data_collections:
                error_string = base_collection_name + " doesn't exist"
                return jsonify({"success": False, "message": error_string, "alert_type": "alert-warning"})

            coll_dict, dm_dict, hl_dict, coll_mdata = user_obj.get_all_collection_info(collection_to_add)
            for dname, doc in coll_dict.items():
                user_obj.append_document_to_collection(base_collection_name, dname, doc, coll_mdata["type"],
                                                       hl_dict[dname], dm_dict[dname])
            user_obj.update_collection_time(base_collection_name)
            self.update_selector_row(self.build_res_dict(base_collection_name, coll_mdata))
            return jsonify({"message": "Collections successfull combined", "alert_type": "alert-success"})
        except Exception as ex:
            return self.get_exception_for_ajax(ex, "Error combining collection")

    def import_as_table(self, collection_name, library_id):
        user_obj = current_user
        file_list = request.files.getlist("file")

        collection_mdata = global_tile_manager.create_initial_metadata()

        known_extensions = [".xlsx", ".csv", ".tsv", ".txt"]

        file_decoding_errors = OrderedDict()
        new_doc_dict = {}
        header_list_dict = {}
        for the_file in file_list:
            filename, file_extension = os.path.splitext(the_file.filename)
            filename = filename.encode("ascii", "ignore")
            if file_extension not in known_extensions:
                return jsonify({"success": False, "message": "Invalid file extension " + file_extension,
                                "alert_type": "alert-warning"})

            decoding_problems = []
            if file_extension == ".xlsx":
                (success, doc_dict, header_dict) = read_excel_file(the_file)
                if not success:  # then doc_dict contains an error object
                    e = doc_dict
                    return jsonify({"message": e["message"], "alert_type": "alert-danger"})
                new_doc_dict.update(doc_dict)
                header_list_dict.update(header_dict)
            else:
                self.show_um_message("Reading file {} and checking encoding".format(filename), library_id,
                                     timeout=10)
                if file_extension == ".csv":
                    (success, row_list, header_list, encoding, decoding_problems) = read_csv_file_to_list(the_file)
                elif file_extension == ".tsv":
                    (success, row_list, header_list, encoding, decoding_problems) = read_tsv_file_to_list(the_file)
                elif file_extension == ".txt":
                    (success, row_list, header_list, encoding, decoding_problems) = read_txt_file_to_list(the_file)
                else:
                    return jsonify({"message": "unkown file extension", "alert_type": "alert-danger"})
                self.show_um_message("Got encoding {} for {}".format(encoding, filename), library_id, timeout=10)

                if not success:  # then row_list contains an error object
                    e = row_list
                    return jsonify({"message": e["message"], "alert_type": "alert-danger"})
                new_doc_dict[filename] = row_list
                header_list_dict[filename] = header_list

            if len(decoding_problems) > 0:
                file_decoding_errors[filename] = decoding_problems

        try:
            user_obj.create_complete_collection(collection_name, new_doc_dict, "table", None,
                                                header_list_dict, collection_mdata)
        except Exception as ex:
            return self.get_exception_for_ajax(ex, "Error creating collection")

        new_row = self.build_res_dict(collection_name, collection_mdata, user_obj)
        if len(file_decoding_errors.keys()) == 0:
            file_decoding_errors = None
        return jsonify({"success": True, "new_row": new_row,
                        "message": "Collection successfully loaded", "alert_type": "alert-success",
                        "file_decoding_errors": file_decoding_errors})

    def import_as_freeform(self, collection_name, library_id):
        user_obj = current_user
        file_list = request.files.getlist("file")

        collection_mdata = global_tile_manager.create_initial_metadata()
        file_decoding_errors = OrderedDict()
        new_doc_dict = {}
        for the_file in file_list:

            filename, file_extension = os.path.splitext(the_file.filename)
            filename = filename.encode("ascii", "ignore")
            self.show_um_message("Reading file {} and checking encoding".format(filename), library_id, timeout=10)
            (success, result_txt, encoding, decoding_problems) = read_freeform_file(the_file)
            self.show_um_message("Got encoding {} for {}".format(encoding, filename), library_id, timeout=10)
            if not success:  # then result_txt contains an error object
                e = result_txt
                return jsonify({"message": e.message, "alert_type": "alert-danger"})
            new_doc_dict[filename] = result_txt
            if len(decoding_problems) > 0:
                file_decoding_errors[filename] = decoding_problems

        try:
            user_obj.create_complete_collection(collection_name, new_doc_dict, "freeform", None,
                                                None, collection_mdata)
        except Exception as ex:
            return self.get_exception_for_ajax(ex, "Error creating collection")

        new_row = self.build_res_dict(collection_name, collection_mdata, user_obj)
        if len(file_decoding_errors.keys()) == 0:
            file_decoding_errors = None
        return jsonify({"success": True, "new_row": new_row,
                        "alert_type": "alert-success",
                        "file_decoding_errors": file_decoding_errors})

    def delete_collection(self):
        try:
            user_obj = current_user
            collection_names = request.json["resource_names"]
            for collection_name in collection_names:
                user_obj.remove_collection(collection_name)
            return jsonify({"success": True, "message": "Collection(s) successfully deleted",
                            "alert_type": "alert-success"})

        except Exception as ex:
            return self.get_exception_for_ajax(ex, "Error deleting collections")

    def combine_to_new_collection(self):
        try:
            user_obj = current_user
            original_collections = request.json["original_collections"]
            new_name = request.json["new_name"]
            coll_dict, dm_dict, hl_dict, coll_mdata = user_obj.get_all_collection_info(original_collections[0])
            user_obj.create_complete_collection(new_name,
                                                coll_dict,
                                                coll_mdata["type"],
                                                dm_dict,
                                                hl_dict,
                                                coll_mdata)
            for col in original_collections[1:]:
                coll_dict, dm_dict, hl_dict, coll_mdata = user_obj.get_all_collection_info(col)
                for dname, doc in coll_dict.items():
                    user_obj.append_document_to_collection(new_name, dname, doc, coll_mdata["type"],
                                                           hl_dict[dname], dm_dict[dname])
            user_obj.update_collection_time(new_name)
            metadata = user_obj.get_collection_metadata(new_name)
            new_row = self.build_res_dict(new_name, metadata, user_obj)
            return jsonify({"success": True, "new_row": new_row})

        except Exception as ex:
            return self.get_exception_for_ajax(ex, "Error combining collections")

    def duplicate_collection(self):
        user_obj = current_user
        new_res_name = request.json["new_res_name"]
        coll_dict, dm_dict, hl_dict, coll_mdata = user_obj.get_all_collection_info(request.json['res_to_copy'])

        result = user_obj.create_complete_collection(new_res_name,
                                                     coll_dict,
                                                     coll_mdata["type"],
                                                     dm_dict,
                                                     hl_dict,
                                                     coll_mdata)

        if not result["success"]:
            result["message"] = result["message"]
            result["alert_type"] = "alert-warning"
            return jsonify(result)

        metadata = user_obj.get_collection_metadata(new_res_name)
        new_row = self.build_res_dict(new_res_name, metadata, user_obj)
        return jsonify({"success": True, "new_row": new_row})


class RepositoryCollectionManager(CollectionManager):
    rep_string = "repository-"
    is_repository = True

    def add_rules(self):
        pass
